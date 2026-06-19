"""
APP INVENTARIO - 100 MONTADITOS LAGOH
======================================
El encargado solo introduce la fecha y el conteo final.
El resto es automático: inventario inicial, Codisys, diferencias y guardado.

Cómo arrancar:
    streamlit run app.py

Después abre en el móvil:  http://[IP-del-PC]:8501
"""

import asyncio
import json
import os
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

# Cargar credenciales — primero Streamlit Secrets (cloud), luego .env (local)
def _cargar_env():
    # 1) Streamlit Secrets (Streamlit Community Cloud)
    try:
        _secrets = st.secrets
        for _k in ["CODISYS_USER", "CODISYS_PASS", "CODISYS_TIENDA"]:
            if _k in _secrets:
                os.environ[_k] = str(_secrets[_k])
        if any(k in _secrets for k in ["CODISYS_USER", "CODISYS_PASS"]):
            return
    except Exception:
        pass
    # 2) Fichero .env local (cuando corre en el PC)
    carpeta = Path(__file__).parent
    for nombre in [".env", ".env.txt", "env.txt", "env"]:
        p = carpeta / nombre
        if p.exists():
            with open(p, encoding="utf-8") as f:
                for linea in f:
                    linea = linea.strip()
                    if linea and not linea.startswith("#") and "=" in linea:
                        clave, valor = linea.split("=", 1)
                        os.environ[clave.strip()] = valor.strip()
            break

_cargar_env()

# Instalar Chromium automáticamente si no está disponible (Streamlit Cloud)
import subprocess as _sp
_sp.run(["playwright", "install", "chromium", "--with-deps"], capture_output=True)

# ═══════════════════════════════════════════════════════════
#  CONFIGURACIÓN DE PRODUCTOS
#  (extraída de tu Excel — no tocar)
# ═══════════════════════════════════════════════════════════

PERFILES = {
    "Laura": {
        "icono": "🥐", "biberones": False,
        "productos": [
            {"nombre": "100M MONTADITO SIN VITOLA 110UD",      "uds_caja": 110,    "precio": 16.78,  "formato": "UNIDAD", "r": 1,     "s": None},
            {"nombre": "EMPOTRAITO 115 UDS + VITOLA",           "uds_caja": 115,    "precio": 19.90,  "formato": "UNIDAD", "r": 1,     "s": None},
            {"nombre": "PAN MONTADITO BRIOCHE 25GR X 192UD",   "uds_caja": 192,    "precio": 54.31,  "formato": "UNIDAD", "r": 16,    "s": None},
            {"nombre": "PAN DINNER ROLL C-126",                 "uds_caja": 126,    "precio": 46.94,  "formato": "UNIDAD", "r": 21,    "s": None},
            {"nombre": "PAN MINI BURGUER 100M 120UD",           "uds_caja": 120,    "precio": 32.75,  "formato": "UNIDAD", "r": 60,    "s": None},
            {"nombre": "PAN MOLLETE C-60UND",                   "uds_caja": 60,     "precio": 12.60,  "formato": "UNIDAD", "r": 1,     "s": None},
            {"nombre": "BARRITA CON CEREALES 70UD",             "uds_caja": 70,     "precio": 25.06,  "formato": "UNIDAD", "r": 1,     "s": None},
            {"nombre": "GILDA DE ANCHOA 4x30UD (120GILDAS)",   "uds_caja": 120,    "precio": 45.63,  "formato": "UNIDAD", "r": 30,    "s": None},
            {"nombre": "GILDA DE BOQUERON 4x30UD (120 GILDAS)","uds_caja": 120,    "precio": 45.63,  "formato": "UNIDAD", "r": 30,    "s": None},
            {"nombre": "CHORIZO PARRILLERO 30Gx20 C-6",        "uds_caja": 120,    "precio": 27.57,  "formato": "UNIDAD", "r": 20,    "s": None},
            {"nombre": "LONCHAS MADURADO 1KGX6U",               "uds_caja": 6,      "precio": 47.71,  "formato": "KILO",   "r": 1,     "s": None},
            {"nombre": "SALCHICHA FRANKFURT 40Gx5 C-40",        "uds_caja": 200,    "precio": 22.66,  "formato": "UNIDAD", "r": 5,     "s": None},
        ],
    },
    "Eli": {
        "icono": "🥫", "biberones": True,
        "productos": [
            {"nombre": "SALSA ALLIOLI CHOVI 768GR X 8UD",          "uds_caja": 6.144,  "precio": 22.44,  "formato": "KILO", "r": 0.768,  "s": None},
            {"nombre": "SALSA BARBACOA HEINZ 2,5KG X 2UD",         "uds_caja": 5,      "precio": 12.11,  "formato": "KILO", "r": 2.5,    "s": 0.39},
            {"nombre": "SALSA BRAVA ORLANDO 1,8KG X 6UD",          "uds_caja": 10.8,   "precio": 28.66,  "formato": "KILO", "r": 1.8,    "s": 0.39},
            {"nombre": "SALSA KETCHUP HEINZ 5,8KG X 3UD HALAL",    "uds_caja": 17.4,   "precio": 28.24,  "formato": "KILO", "r": 5.8,    "s": 0.39},
            {"nombre": "SALSA MAYONESA CHOVI 3,494KG X 4UD",       "uds_caja": 13.976, "precio": 39.11,  "formato": "KILO", "r": 3.494,  "s": 0.39},
            {"nombre": "SALSA MOJO PICON CUBO 900GR X 4UD",        "uds_caja": 3,      "precio": 19.62,  "formato": "KILO", "r": 0.9,    "s": 0.39},
            {"nombre": "SALSA MOSTAZA & MIEL CHOVI 855GR X 8UD",   "uds_caja": 6.84,   "precio": 27.11,  "formato": "KILO", "r": 0.8553, "s": None},
            {"nombre": "SALSA 100M 0,855GR X 8UD",                  "uds_caja": 6.84,   "precio": 35.60,  "formato": "KILO", "r": 0.855,  "s": None},
            {"nombre": "SALSA CHEDDAR BOLSA 1KG X 6UD",             "uds_caja": 6,      "precio": 19.56,  "formato": "KILO", "r": 1.0,    "s": 0.39},
            {"nombre": "GUACAMOLE MEDITERRANEO 0,5KG X 8UD",        "uds_caja": 4,      "precio": 20.54,  "formato": "KILO", "r": 0.5,    "s": 0.39},
            {"nombre": "MERMELADA DE FRESA BIB 2,75KG X 3UD",       "uds_caja": 8.25,   "precio": 28.54,  "formato": "KILO", "r": 2.75,   "s": None},
            {"nombre": "SALSA PIZZA 1KG X 10UD",                     "uds_caja": 10,     "precio": 53.81,  "formato": "KILO", "r": 1.0,    "s": 0.39},
            {"nombre": "MOZZARELLA CUBETTI 1KG C-10",                "uds_caja": 10,     "precio": 54.92,  "formato": "KILO", "r": 1.0,    "s": None},
            {"nombre": "LOMO AL AJILLO 500GR X 10UD",                "uds_caja": 5,      "precio": 27.18,  "formato": "KILO", "r": 0.5,    "s": None},
            {"nombre": "PEPPERONI LONCHA 6X500GR",                   "uds_caja": 3,      "precio": 24.46,  "formato": "KILO", "r": 0.5,    "s": None},
        ],
    },
    "Marta": {
        "icono": "🍗", "biberones": False,
        "productos": [
            {"nombre": "ALITA BARBACOA 2KG (C-2 UN)",               "uds_caja": 4,      "precio": 19.95,  "formato": "KILO", "r": 2,     "s": None},
            {"nombre": "CARRILLERAS ASADAS 0,8KG 8UD",              "uds_caja": 6.4,    "precio": 60.76,  "formato": "KILO", "r": 0.8,   "s": None},
            {"nombre": "CROQUETA JAMON 1KG X 5UD",                  "uds_caja": 5,      "precio": 19.01,  "formato": "KILO", "r": 1,     "s": None},
            {"nombre": "LAGRIMAS DE POLLO (C-5 KG)",                 "uds_caja": 5,      "precio": 31.41,  "formato": "KILO", "r": 1,     "s": None},
            {"nombre": "MINI CALAMAR EN TEMPURA VGAMA 2KG",           "uds_caja": 2,      "precio": 14.58,  "formato": "KILO", "r": 2,     "s": None},
            {"nombre": "PALOMITA GOUDA 1KG X 5UD",                   "uds_caja": 5,      "precio": 22.62,  "formato": "KILO", "r": 1,     "s": None},
            {"nombre": "PALOMITA POLLO CJ-5KG",                      "uds_caja": 5,      "precio": 20.86,  "formato": "KILO", "r": 1,     "s": None},
            {"nombre": "PATATA PRIVATE RVA 6X6 2.5KX4U",              "uds_caja": 10,     "precio": 17.08,  "formato": "KILO", "r": 2.5,   "s": None},
            {"nombre": "POLLO ASADO KEBAB TIRAS HALAL 1KG X 6UD",   "uds_caja": 6,      "precio": 32.93,  "formato": "KILO", "r": 1,     "s": None},
            {"nombre": "POLLO ASADO TIRAS HALAL 1KG X 6UD",         "uds_caja": 6,      "precio": 33.23,  "formato": "KILO", "r": 1,     "s": None},
            {"nombre": "SALMON AHUMADO CONGELADO 500GR X 12UD",     "uds_caja": 6,      "precio": 126.72, "formato": "KILO", "r": 0.5,   "s": None},
            {"nombre": "TORTILLA PL. C/CEBO 750GR 12UD",             "uds_caja": 9,      "precio": 29.35,  "formato": "KILO", "r": 0.75,  "s": None},
        ],
    },
    "David": {
        "icono": "🛒", "biberones": False, "sin_conteo": True,
        "productos": [
            # ── PAN ──────────────────────────────────────────────────────────
            {"nombre": "100M MONTADITO SIN VITOLA 110UD",                   "uds_caja": 110,     "precio": 16.72,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "EMPOTRAITO 115 UDS + VITOLA",                        "uds_caja": 115,     "precio": 19.95,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "PAN MONTADITO BRIOCHE 25GR X 192UD",                 "uds_caja": 192,     "precio": 52.33,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "PAN DINNER ROLL C-126",                              "uds_caja": 126,     "precio": 46.94,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "PAN MINI BURGUER 100M 120UD",                        "uds_caja": 120,     "precio": 32.69,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "PAN MOLLETE C-60UND",                                "uds_caja": 60,      "precio": 12.51,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "BARRITA CON CEREALES 70UD",                          "uds_caja": 70,      "precio": 24.49,   "formato": "UNIDAD", "r": 1, "s": None},
            # ── CONGELADO ────────────────────────────────────────────────────
            {"nombre": "ALITA BARBACOA 2 KG (C-2 UN)",                       "uds_caja": 4,       "precio": 19.88,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "CARRILLERAS ASADAS 0,8 KG 8UD",                      "uds_caja": 6.4,     "precio": 59.92,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "CONO MENU VAINILLA-CHOCOLATE 40UD",                  "uds_caja": 40,      "precio": 18.99,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "CROQUETA JAMON IB.  1KG X 5UN",                      "uds_caja": 5,       "precio": 19.55,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "DOMINO VAINILLA 90ML X 42UD",                        "uds_caja": 42,      "precio": 17.26,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "GANCHITO DE CHEETOS 1KG X 4UD",                      "uds_caja": 4,       "precio": 33.61,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "GUACAMOLE MEDITERRANEO 0,5KG X 8UD",                 "uds_caja": 4,       "precio": 20.47,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "LAGRIMAS DE POLLO (C-5 KG)",                         "uds_caja": 5,       "precio": 31.33,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "MINI CALAMAR EN TEMPURA VGAMA 2KG",                  "uds_caja": 2,       "precio": 14.27,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "MINI HAMBURGUESA MIXTA 1,375KGX2(110UD)",            "uds_caja": 2.75,    "precio": 23.79,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PALOMITA GOUDA 1 KG X 5 UD",                         "uds_caja": 5,       "precio": 22.13,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PALOMITA POLLO CJ- 5 KG",                            "uds_caja": 5,       "precio": 20.41,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PATATA PRIVATE RVA 6X6 2.5KX4U",                     "uds_caja": 10,      "precio": 16.66,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "POLLO ASADO KEBAB TIRAS HALAL 1KG X 6UD",            "uds_caja": 6,       "precio": 32.82,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "POLLO ASADO TIRAS HALAL 1KG X 6UD",                  "uds_caja": 6,       "precio": 33.12,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PULLED PORK 500GR X 10UD",                           "uds_caja": 5,       "precio": 33.93,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALMON AHUMADO CONGELADO 500GR X 12UD (6KG APR",     "uds_caja": 6,       "precio": 124.18,  "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "TORTILLA PL. C/CEBO 750gr 12ud",                     "uds_caja": 9,       "precio": 29.2,    "formato": "KILO",   "r": 1, "s": None},
            # ── REFRIGERADO ──────────────────────────────────────────────────
            {"nombre": "ACEITUNA ABUELA GALON2,5NEX2U",                      "uds_caja": 5,       "precio": 17.88,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "BACON FRITO TIRAS 1KG X 4UD",                        "uds_caja": 4,       "precio": 30.57,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "GILDA DE ANCHOA 4x30UD (120GILDAS)",                 "uds_caja": 120,     "precio": 45.63,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "GILDA DE BOQUERON 4x30UD (120 GILDAS)",              "uds_caja": 120,     "precio": 45.63,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "JAMON SERRANO GRAN RVA LONCH 1.2KG C-8",             "uds_caja": 9.6,     "precio": 106.83,  "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "LOMO AL AJILLO 500GR X 10UD",                        "uds_caja": 5,       "precio": 27.09,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "LONCHAS MADURADO 1KGX6U",                            "uds_caja": 6,       "precio": 46.72,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "MOZZARELLA CUBETTI 1KG C-10",                        "uds_caja": 10,      "precio": 54.92,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PEPPERONI LONCHA  6X500GR",                          "uds_caja": 3,       "precio": 24.36,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA CHEDDAR BOLSA 1KG X 6 UD",                     "uds_caja": 6,       "precio": 19.46,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "TORREZNO DADOS 1KG X 5UD",                           "uds_caja": 5,       "precio": 50.39,   "formato": "KILO",   "r": 1, "s": None},
            # ── SECO ─────────────────────────────────────────────────────────
            {"nombre": "ACEITE ESPECIAL FRITURA 5LX3UD",                     "uds_caja": 15,      "precio": 33.04,   "formato": "LITRO",  "r": 1, "s": None},
            {"nombre": "ACEITE OVE TARRINA 13ml 200ud",                      "uds_caja": 200,     "precio": 18.03,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "ATUN EN ACEITE DE GIRASOL 1KGX16U",                  "uds_caja": 16,      "precio": 96.26,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "CEBOLLA FRITA 1KG X 10UD",                           "uds_caja": 10,      "precio": 43.97,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "DORITOS NATUREL 455GR X 12UD",                       "uds_caja": 5.46,    "precio": 22.94,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "MERMELADA DE FRESA BIB 2,75KG X 3UD",                "uds_caja": 8.25,    "precio": 28.04,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PATATA PAJA PALA (C-12UN X 0,365 KG)",               "uds_caja": 4.38,    "precio": 19.01,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "PATATAS LAY'S 480GR X 9UD",                         "uds_caja": 4.32,    "precio": 17.09,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA 100M 0,855GR X 8UD",                           "uds_caja": 6.84,    "precio": 40.6,    "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA ALLIOLI CHOVI 768,25GR X 8UD",                 "uds_caja": 6.144,   "precio": 22.33,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA BARBACOA HEINZ 2,5KG X 2UD",                   "uds_caja": 5,       "precio": 12.11,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA BRAVA ORLANDO 1,8KG X 6UD",                    "uds_caja": 10.8,    "precio": 28.66,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA KETCHUP HEINZ 5,8KG X 3UD HALAL",              "uds_caja": 17.4,    "precio": 27.8,    "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA MAYONESA CHOVI CUBO 3,494KG X 4UD",            "uds_caja": 13.976,  "precio": 38.87,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA MOJO PICON CUBO 900GR X 4UD",                  "uds_caja": 3,       "precio": 19.52,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "SALSA MOSTAZA & MIEL CHOVI 855,3GR X 8UD",           "uds_caja": 6.84,    "precio": 26.99,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "TOMAUNTA 800 GR. C-12UD",                            "uds_caja": 9.6,     "precio": 16.51,   "formato": "KILO",   "r": 1, "s": None},
            {"nombre": "TORTILLAS TRIGO 18UD 24P 12CM",                      "uds_caja": 432,     "precio": 40.09,   "formato": "UNIDAD", "r": 1, "s": None},
            # ── BOLLERÍA / DESAYUNOS ─────────────────────────────────────────
            {"nombre": "CROISSANT MANTEQUILLA 55GR 30U",                     "uds_caja": 30,      "precio": 10.41,   "formato": "UNIDAD", "r": 1, "s": None},
            {"nombre": "MANTEQUILLA 10 G X 100 UD",                          "uds_caja": 100,     "precio": 9.43,    "formato": "UNIDAD", "r": 1, "s": None},
        ],
    },
}
# Placeholder — sobreescrito tras seleccion de perfil
PRODUCTOS       = PERFILES["Laura"]["productos"]
TIENE_BIBERONES = False

CODISYS_URL   = "https://www.codisysdc.com/Account/Login?ReturnUrl=%2F"
CODISYS_USER  = os.getenv("CODISYS_USER", "")
CODISYS_PASS  = os.getenv("CODISYS_PASS", "")
CODISYS_TIENDA = os.getenv("CODISYS_TIENDA", "100M BLUE CC LAGOH")

STATE_FILE    = Path("inventario_estado.json")
HISTORIAL_FILE = Path("inventario_historial.json")


# ═══════════════════════════════════════════════════════════
#  ESTADO — inventario inicial = final del último periodo
# ═══════════════════════════════════════════════════════════

def cargar_estado() -> dict:
    """Devuelve el inventario final guardado del último periodo."""
    if STATE_FILE.exists():
        with open(STATE_FILE, encoding="utf-8") as f:
            return json.load(f)
    # Primera vez: inventario inicial en cero
    return {p["nombre"]: {"final": 0.0, "fecha": None} for p in PRODUCTOS}

def guardar_estado(resultados: list, fecha: str):
    """Guarda el inventario final para que sea el inicial del próximo periodo."""
    estado = {}
    for r in resultados:
        estado[r["nombre"]] = {"final": r["inv_final"], "fecha": fecha}
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(estado, f, ensure_ascii=False, indent=2)

def cargar_historial() -> list:
    if HISTORIAL_FILE.exists():
        with open(HISTORIAL_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []

def guardar_historial(entrada: dict):
    historial = cargar_historial()
    historial.append(entrada)
    with open(HISTORIAL_FILE, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)

def borrar_historial_idx(idx: int):
    historial = cargar_historial()
    if 0 <= idx < len(historial):
        historial.pop(idx)
        with open(HISTORIAL_FILE, "w", encoding="utf-8") as f:
            json.dump(historial, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════
#  CÁLCULOS
# ═══════════════════════════════════════════════════════════

def calcular_inv_final(prod: dict, cajas: float, paquetes: float, sueltos: float, biberones: float = 0.0) -> float:
    """Inventario final = cajas × uds_caja + paquetes × r + biberones × s + sueltos"""
    s = prod.get("s") or 0.0
    return (cajas * prod["uds_caja"]) + (paquetes * prod["r"]) + (biberones * s) + sueltos

def calcular_fila(prod: dict, inv_inicial: float, compras: float,
                  cajas: float, paquetes: float, sueltos: float,
                  usado_codisys: float, ventas_netas: float = 0.0) -> dict:
    coste_ud = prod["precio"] / prod["uds_caja"]
    inv_final = calcular_inv_final(prod, cajas, paquetes, sueltos)
    usado_inv = inv_inicial + compras - inv_final
    diferencia = usado_codisys - usado_inv if usado_codisys is not None else None
    dif_eur = (diferencia * coste_ud) if diferencia is not None else None

    return {
        "nombre":        prod["nombre"],
        "inv_inicial":   inv_inicial,
        "compras":       compras,
        "cajas":         cajas,
        "paquetes":      paquetes,
        "sueltos":       sueltos,
        "inv_final":     inv_final,
        "usado_inv":     usado_inv,
        "usado_cod":     usado_codisys,
        "diferencia":    diferencia,
        "dif_eur":       dif_eur,
        "coste_ud":      coste_ud,
        "ventas_netas":  ventas_netas,
    }


# ═══════════════════════════════════════════════════════════
#  CODISYS — descarga automática
# ═══════════════════════════════════════════════════════════

async def _js_click_btn(page, btn_idx: int):
    """Click directo en .dx-dropdowneditor-button[idx] via JS (sin checks de visibilidad)."""
    await page.evaluate(f"""
        () => {{
            const btns = document.querySelectorAll('.dx-dropdowneditor-button');
            if (btns[{btn_idx}]) btns[{btn_idx}].click();
        }}
    """)


async def _popup_seleccionar(page, opcion: str = "") -> list:
    """Selecciona un item del popup DevExpress abierto.
    Solo considera popups realmente visibles en pantalla."""
    await page.wait_for_timeout(1200)
    result = await page.evaluate("""
        () => {
            const SELECTORS = [
                '.dx-dropdownlist-popup-wrapper',
                '.dx-popup-content',
                '.dx-overlay-content',
                '.dx-selectbox-popup-wrapper'
            ];
            const vp = { w: window.innerWidth, h: window.innerHeight };
            for (const sel of SELECTORS) {
                for (const container of document.querySelectorAll(sel)) {
                    const r = container.getBoundingClientRect();
                    if (r.width <= 0 || r.height <= 0) continue;
                    if (r.x < -10 || r.y < -10 || r.x > vp.w || r.y > vp.h) continue;
                    const st = window.getComputedStyle(container);
                    if (st.display === 'none' || st.visibility === 'hidden' || parseFloat(st.opacity) < 0.1) continue;
                    const listItems = Array.from(container.querySelectorAll('.dx-list-item'));
                    if (listItems.length === 0) continue;
                    return { sel, items: listItems.map((el, i) => ({ i, text: el.textContent.trim() })) };
                }
            }
            return null;
        }
    """)
    if not result:
        print(f"DEBUG: popup vacío")
        return []

    item_list = result['items']
    print(f"DEBUG: popup items = {[x['text'] for x in item_list[:8]]}")

    target_i = 0
    if opcion:
        for item in item_list:
            if opcion.lower() in item['text'].lower():
                target_i = item['i']
                print(f"DEBUG: encontrado '{item['text']}'")
                break

    await page.evaluate(f"""
        () => {{
            const SELECTORS = ['.dx-dropdownlist-popup-wrapper', '.dx-popup-content',
                               '.dx-overlay-content', '.dx-selectbox-popup-wrapper'];
            const vp = {{ w: window.innerWidth, h: window.innerHeight }};
            for (const sel of SELECTORS) {{
                for (const container of document.querySelectorAll(sel)) {{
                    const r = container.getBoundingClientRect();
                    if (r.width <= 0 || r.height <= 0) continue;
                    if (r.x < -10 || r.y < -10 || r.x > vp.w || r.y > vp.h) continue;
                    const st = window.getComputedStyle(container);
                    if (st.display === 'none' || st.visibility === 'hidden') continue;
                    const items = container.querySelectorAll('.dx-list-item');
                    if (items.length === 0) continue;
                    if (items[{target_i}]) items[{target_i}].click();
                    return;
                }}
            }}
        }}
    """)
    await page.wait_for_timeout(500)
    return item_list


async def _descargar_excel_reporte(page, nombre_reporte: str, desde: str, hasta: str, tmp: Path, nombre_fichero: str, familia: str = "", tienda_es_tagbox: bool = False, url_directa: str = "", expandir_secciones: bool = False) -> Path | None:
    """Navega a un informe de Codisys, rellena el formulario y descarga el Excel.

    Estrategia:
      - Todos los clicks en dropdowns del formulario se hacen via JavaScript (.click())
        para evitar problemas de visibilidad de Playwright.
      - Los botones de parámetro se identifican por posición visual (x > mitad viewport,
        ordenados por Y): [0]=Tienda [1]=Desde [2]=Hasta [3]=Familias [4]=Productos.
      - La exportación a Excel también se hace via JS, escaneando todos los nodos de texto
        visibles en busca de "XLSx" / "Excel".
    """
    import unicodedata

    try:
        # ── 0. Navegar al informe ────────────────────────────────────────
        await page.goto("https://www.codisysdc.com/Intermedia_Informes")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)

        if expandir_secciones:
            # Scroll para forzar carga de secciones colapsadas (solo necesario para informes fuera de FAVORITOS)
            for _ in range(4):
                await page.evaluate("window.scrollBy(0, 400)")
                await page.wait_for_timeout(600)
            await page.wait_for_timeout(1000)

            secciones_expandidas = await page.evaluate("""
                () => {
                    function norm(s) {
                        return s.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '');
                    }
                    const keywords = ['informes de compra', 'informes de venta', 'informes de ventas', 'informes de analisis', 'otros'];
                    const expandidos = [];
                    for (const el of document.querySelectorAll('div,span,p,a,li,button,h2,h3,h4')) {
                        const t = norm((el.innerText||el.textContent||'').trim());
                        if (t.length < 4 || t.length > 50) continue;
                        for (const kw of keywords) {
                            if (t === kw) {
                                el.click();
                                expandidos.push(t);
                                break;
                            }
                        }
                    }
                    return expandidos;
                }
            """)
            print(f"DEBUG: secciones expandidas = {secciones_expandidas}")
            await page.wait_for_timeout(2000)

            # Scroll adicional para cargar contenido expandido
            for _ in range(3):
                await page.evaluate("window.scrollBy(0, 400)")
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(1000)

            # Volver arriba para que el click de Playwright funcione bien
            await page.evaluate("window.scrollTo(0, 0)")
            await page.wait_for_timeout(800)

        sin_tilde = ''.join(
            c for c in unicodedata.normalize('NFD', nombre_reporte)
            if unicodedata.category(c) != 'Mn'
        )

        clicked = False
        # Intentos Playwright con texto exacto y sin tilde (NO primera palabra sola — demasiado genérica)
        for intento in [nombre_reporte, sin_tilde]:
            try:
                await page.click(f'text={intento}', timeout=3000)
                clicked = True
                print(f"DEBUG: click en informe '{intento}'")
                break
            except Exception as e:
                err_lower = str(e).lower()
                # "context was destroyed" significa que el click SÍ funcionó y causó navegación
                if "context was destroyed" in err_lower or "navigation" in err_lower:
                    clicked = True
                    print(f"DEBUG: click en informe '{intento}' (navegó — contexto destruido)")
                    break

        # Fallback JS: scan completo del DOM normalizado
        if not clicked:
            def _norm(s):
                return ''.join(
                    c for c in unicodedata.normalize('NFD', s.lower())
                    if unicodedata.category(c) != 'Mn'
                )
            objetivo = _norm(nombre_reporte)
            parcial  = _norm(" ".join(nombre_reporte.split()[:2]))
            try:
                resultado = await page.evaluate(f"""
                    () => {{
                        function norm(s) {{
                            return s.toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '');
                        }}
                        const obj = "{objetivo}";
                        const par = "{parcial}";
                        for (const el of document.querySelectorAll('div,span,p,a,li,td,button')) {{
                            const t = norm((el.textContent||'').trim());
                            if (t.length < 3 || t.length > 80) continue;
                            if (t === obj || t.includes(par)) {{
                                el.click();
                                return 'OK:' + (el.textContent||'').trim();
                            }}
                        }}
                        const textos = [];
                        document.querySelectorAll('div,span').forEach(el => {{
                            const t = (el.textContent||'').trim();
                            if (t.length > 4 && t.length < 60 && !textos.includes(t)) textos.push(t);
                        }});
                        return 'NO:' + textos.slice(0,50).join('|');
                    }}
                """)
                if str(resultado).startswith('OK:'):
                    clicked = True
                    print(f"DEBUG: click JS ({resultado})")
                else:
                    print(f"DEBUG: no encontrado. DOM: {resultado}")
            except Exception as e:
                err_lower = str(e).lower()
                if "context was destroyed" in err_lower or "navigation" in err_lower:
                    clicked = True
                    print(f"DEBUG: click JS causó navegación (contexto destruido)")
                else:
                    raise

        if not clicked:
            raise Exception(f"No se encontró el informe '{nombre_reporte}'")

        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(1500)
        # Esperar a que los botones DevExpress aparezcan (hasta 15s)
        for _w in range(30):
            try:
                _cnt = await page.evaluate("() => document.querySelectorAll('.dx-dropdowneditor-button').length")
            except Exception as _ev_err:
                # La página aún está navegando — esperar y reintentar
                if "context was destroyed" in str(_ev_err).lower() or "navigation" in str(_ev_err).lower():
                    await page.wait_for_load_state("networkidle")
                    await page.wait_for_timeout(1000)
                    continue
                raise
            if _cnt > 0:
                break
            await page.wait_for_timeout(500)
        await page.wait_for_timeout(500)
        print(f"DEBUG: URL formulario = {page.url}")

        # ── 1. Detectar botones de parámetro por posición visual ────────
        # Obtener BoundingClientRect de todos los .dx-dropdowneditor-button visibles
        btns_pos = await page.evaluate("""
            () => {
                const btns = Array.from(document.querySelectorAll('.dx-dropdowneditor-button'));
                return btns.map((btn, i) => {
                    const r = btn.getBoundingClientRect();
                    return { i, x: r.x, y: r.y, w: r.width, h: r.height };
                }).filter(b => b.w > 0 && b.h > 0);
            }
        """)
        print(f"DEBUG: {len(btns_pos)} botones dropdown visibles")
        for b in btns_pos:
            print(f"  [{b['i']}] x={b['x']:.0f} y={b['y']:.0f}")

        # Filtrar los que están a la derecha de la pantalla (panel de parámetros)
        # y ordenar por Y para obtener el orden visual: Tienda, Desde, Hasta, Familias, Productos
        mid_x = 640  # viewport 1280px
        param_btns = sorted([b for b in btns_pos if b['x'] > mid_x], key=lambda b: b['y'])
        if len(param_btns) < 3:
            # Fallback: si el panel está a la izquierda, usar todos ordenados por Y
            param_btns = sorted(btns_pos, key=lambda b: b['y'])
        print(f"DEBUG: {len(param_btns)} botones de parámetro (ordenados por Y)")

        # Helper: click tagbox nth(n), seleccionar item o "Seleccionar Todos", cerrar popup
        async def _tagbox_seleccionar(nth: int, opcion: str = "", seleccionar_todos: bool = False):
            tagboxes = page.locator('.custom-multivalue')
            await tagboxes.nth(nth).click(force=True)
            await page.wait_for_timeout(1200)
            if seleccionar_todos:
                sel_ok = await page.evaluate("""
                    () => {
                        const vp = { w: window.innerWidth, h: window.innerHeight };
                        for (const sel of ['.dx-list-select-all', '[class*="select-all"]', '.dx-checkbox']) {
                            for (const el of document.querySelectorAll(sel)) {
                                const r = el.getBoundingClientRect();
                                if (r.width <= 0 || r.height <= 0) continue;
                                if (r.x < -10 || r.y < -10 || r.x > vp.w || r.y > vp.h) continue;
                                el.click();
                                return true;
                            }
                        }
                        return false;
                    }
                """)
                print(f"DEBUG: tagbox[{nth}] Seleccionar todos = {sel_ok}")
            else:
                await _popup_seleccionar(page, opcion)
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)

        # ── 2. Tienda ────────────────────────────────────────────────────
        if tienda_es_tagbox:
            # Tiendas es tagbox → nth(0), seleccionar todos (solo hay 1 tienda)
            print("DEBUG: Tienda → tagbox nth(0) (Seleccionar Todos)")
            await _tagbox_seleccionar(0, seleccionar_todos=True)
        else:
            # Tiendas es dropdown estándar
            if param_btns:
                t_idx = param_btns[0]['i']
                print(f"DEBUG: Tienda → btn[{t_idx}] (JS click)")
                await _js_click_btn(page, t_idx)
                await _popup_seleccionar(page, CODISYS_TIENDA)
                await page.wait_for_timeout(2000)

        # ── 3. Fechas ────────────────────────────────────────────────────
        # Cerrar cualquier popup/overlay abierto antes de interactuar con fechas
        await page.evaluate("""
            () => {
                // Cerrar overlays DevExpress abiertos
                const overlays = document.querySelectorAll('.dx-overlay-wrapper, .dx-popup-wrapper');
                overlays.forEach(el => {
                    try { el.style.display = 'none'; } catch(e) {}
                });
            }
        """)
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(400)
        await page.mouse.click(640, 60)
        await page.wait_for_timeout(300)

        date_inputs = page.locator('.dx-datebox .dx-texteditor-input')
        cnt_dt = await date_inputs.count()
        print(f"DEBUG: inputs de fecha = {cnt_dt}")
        if cnt_dt >= 2:
            for idx_f, val in [(0, desde), (1, hasta)]:
                fi = date_inputs.nth(idx_f)
                await fi.scroll_into_view_if_needed()
                await fi.click(force=True)
                await page.wait_for_timeout(200)
                await page.keyboard.press("Control+a")
                await page.keyboard.type(val)
                await page.keyboard.press("Tab")
                await page.wait_for_timeout(400)
                actual = await fi.input_value()
                print(f"DEBUG: fecha[{idx_f}] enviado='{val}' resultado='{actual}'")

        # ── 4. Familias y Productos (solo si existen tagboxes suficientes) ─
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)

        cnt_tagboxes = await page.locator('.custom-multivalue').count()
        print(f"DEBUG: total tagboxes en formulario = {cnt_tagboxes}")

        fam_nth = 1 if tienda_es_tagbox else 0
        prod_nth = 2 if tienda_es_tagbox else 1

        if cnt_tagboxes > fam_nth:
            await _tagbox_seleccionar(fam_nth, opcion=familia)
        else:
            print(f"DEBUG: sin tagbox Familias (cnt={cnt_tagboxes}, fam_nth={fam_nth}) — omitido")

        if cnt_tagboxes > prod_nth:
            await _tagbox_seleccionar(prod_nth, seleccionar_todos=True)
        else:
            print(f"DEBUG: sin tagbox Productos (cnt={cnt_tagboxes}, prod_nth={prod_nth}) — omitido")

        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)

        # ── 6. VER INFORME ───────────────────────────────────────────────
        print("DEBUG: click VER INFORME")
        await page.evaluate("""
            () => {
                const btns = Array.from(document.querySelectorAll('.dx-button, button, [class*="dx-button"]'));
                const ver = btns.find(b => b.textContent.trim().toLowerCase().includes('ver informe'));
                if (ver) ver.click();
            }
        """)
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)
        print(f"DEBUG: URL tras VER INFORME = {page.url}")

        # ── 7. Esperar botón Exportar y que el informe esté listo ────────────────
        # Esperar a que el informe esté listo (hasta 40s) y descargar Excel
        xlsx_path = tmp / nombre_fichero
        descargado = False
        for _intento in range(8):   # cada 5s, hasta ~40s
            await page.wait_for_timeout(5000)
            # ¿Existe algún botón de exportar?
            hay_export = await page.evaluate("""
                () => {
                    const sels = ['[title="Exportar"]', '[title="Export"]'];
                    for (const s of sels) {
                        if (document.querySelectorAll(s).length) return true;
                    }
                    return false;
                }
            """)
            print(f"DEBUG: intento {_intento+1} — botón exportar: {hay_export}")
            if not hay_export:
                continue

            # Estrategia: abrir listener de descarga, luego click en Exportar
            # y si aparece submenú, click en XLSX dentro del mismo listener
            try:
                async with page.expect_download(timeout=12000) as dl:
                    # Click en botón Exportar
                    await page.evaluate("""
                        () => {
                            for (const s of ['[title="Exportar"]', '[title="Export"]']) {
                                const btns = [...document.querySelectorAll(s)]
                                    .filter(el => el.getBoundingClientRect().width > 0);
                                if (btns.length) { btns[btns.length-1].click(); return; }
                            }
                        }
                    """)
                    # Esperar brevemente por si aparece submenú de formatos
                    await page.wait_for_timeout(1000)
                    # Listar items visibles del posible submenú
                    items = await page.evaluate("""
                        () => {
                            const sels = [
                                '.dx-list-item', '.dx-menu-item', '.dx-item',
                                'li', '[role="menuitem"]', '[role="option"]',
                                '.dx-popup-content li', '.dx-popup-content .dx-item'
                            ];
                            const res = [];
                            for (const s of sels) {
                                for (const el of document.querySelectorAll(s)) {
                                    const r = el.getBoundingClientRect();
                                    const t = (el.textContent||'').trim();
                                    if (r.width > 0 && r.height > 0 && t) res.push(t);
                                }
                            }
                            return [...new Set(res)].slice(0, 15);
                        }
                    """)
                    print(f"DEBUG: items visibles tras click exportar = {items}")
                    # Si hay items de submenú, hacer click en XLSX/Excel
                    if items:
                        await page.evaluate("""
                            () => {
                                const kws = ['xlsx', 'excel'];
                                const sels = [
                                    '.dx-list-item', '.dx-menu-item', '.dx-item',
                                    'li', '[role="menuitem"]', '[role="option"]',
                                    '.dx-popup-content li', '.dx-popup-content .dx-item'
                                ];
                                for (const s of sels) {
                                    for (const el of document.querySelectorAll(s)) {
                                        const t = (el.textContent||'').trim().toLowerCase();
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0 && r.height > 0 && kws.some(k => t.includes(k))) {
                                            el.click();
                                            return;
                                        }
                                    }
                                }
                            }
                        """)
                        await page.wait_for_timeout(500)
                download = await dl.value
                await download.save_as(xlsx_path)
                descargado = True
                print(f"DEBUG: Excel descargado ({nombre_reporte}) en intento {_intento+1}")
                break
            except Exception as _dl_err:
                print(f"DEBUG: descarga intento {_intento+1} falló: {str(_dl_err)[:120]}")
        if not descargado:
            raise Exception(f"No se pudo descargar '{nombre_reporte}' tras 40s")

        return xlsx_path

    except Exception as e:
        import traceback
        print(f"Error descargando '{nombre_reporte}': {traceback.format_exc()}")
        return None


def _parsear_consumos(xlsx_path: Path) -> dict:
    """Parsea el Excel de Consumo Materias Primas Venta.
    Estructura real:
      Fila 1 = cabecera: DESTIENDAS, DESFAMILIAS, CODIGO, DESCRIPCION,
                          FAM_COD, FAM_DES, UND_MEDIDA, UND_PRE, CAN_VENTA, CAN_DESPERDICIO, ...
      Columna producto : DESCRIPCION
      Columna consumo  : CAN_VENTA (cantidad consumida/vendida)
    Devuelve: {DESCRIPCION: valor} y también guarda _cod_map={CODIGO: DESCRIPCION} en módulo
    """
    global _codisys_cod_map, _codisys_cod_val_map
    datos = {}
    _codisys_cod_map = {}
    _codisys_cod_val_map = {}
    try:
        df = pd.read_excel(xlsx_path)          # fila 1 ya es la cabecera
        df.columns = [str(c).strip().upper() for c in df.columns]
        print(f"DEBUG consumos columnas: {list(df.columns)}")

        prod_col  = next((c for c in df.columns if c == "DESCRIPCION"), None)
        cod_col   = next((c for c in df.columns if c == "CODIGO"), None)
        # Prioridad: CAN_TOTAL (Total Consumo) > cualquier col con TOTAL > CAN_VENTA
        valor_col = next((c for c in df.columns if c == "CAN_TOTAL"), None)
        if valor_col is None:
            valor_col = next((c for c in df.columns if "TOTAL" in c), None)
        if valor_col is None:
            valor_col = next((c for c in df.columns if c == "CAN_VENTA"), None)

        print(f"DEBUG consumos: prod_col={prod_col}  cod_col={cod_col}  valor_col={valor_col}")

        if prod_col and valor_col:
            _cols = [c for c in [cod_col, prod_col, valor_col] if c]
            df = df[_cols].copy()
            df[prod_col]  = df[prod_col].astype(str).str.strip().str.rstrip(",")
            df[valor_col] = pd.to_numeric(df[valor_col], errors="coerce").fillna(0)
            df = df[df[prod_col].str.len() > 2]
            df = df[~df[prod_col].str.upper().isin(["DESCRIPCION", "TOTAL", "NAN", ""])]
            # Mapa CODIGO → DESCRIPCION para matching exacto por código
            if cod_col:
                for _, r in df.iterrows():
                    cod = str(r[cod_col]).strip()
                    if cod and cod.upper() not in ("NAN", ""):
                        _codisys_cod_map[cod] = str(r[prod_col]).strip()
            # Agrupar por producto (por si aparece varias veces)
            resumen = df.groupby(prod_col)[valor_col].sum()
            datos = {k: float(v) for k, v in resumen.items()}
            # También indexar por código para matching exacto
            if cod_col:
                _cod_val = df.groupby(cod_col)[valor_col].sum()
                for _cod_k, _cod_v in _cod_val.items():
                    _codisys_cod_val_map[str(_cod_k).strip()] = float(_cod_v)
            print(f"DEBUG consumos: {len(datos)} productos, {len(_codisys_cod_map)} códigos")
        else:
            print(f"DEBUG consumos: columnas no encontradas. Disponibles: {list(df.columns)}")
    except Exception as e:
        import traceback
        print(f"Error parseando consumos: {traceback.format_exc()}")
    return datos

_codisys_cod_map: dict = {}      # CODIGO → DESCRIPCION, poblado por _parsear_consumos
_codisys_cod_val_map: dict = {}  # CODIGO → valor consumo, poblado por _parsear_consumos

# Mapa producto_nombre → cod_codisys, leído de cod_codisys_map.json
def _cargar_cod_codisys_map() -> dict:
    """Lee cod_codisys_map.json y devuelve {nombre_producto: codigo_codisys}."""
    import json
    p = Path(__file__).parent / "cod_codisys_map.json"
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}

COD_CODISYS_MAP: dict = _cargar_cod_codisys_map()  # {nombre_perfiles: cod_codisys}


_codisys_compras_cod_map: dict = {}   # CODIGO (sin ERE) → total cantidad comprada

def _parsear_compras(xlsx_path: Path) -> dict:
    """Parsea el Excel de Mercancía Recibida.
    Devuelve {nombre_producto: total_cantidad}.
    También rellena _codisys_compras_cod_map {codigo: total_cantidad} para matching exacto.
    """
    global _codisys_compras_cod_map
    _codisys_compras_cod_map = {}
    datos = {}
    try:
        df = pd.read_excel(xlsx_path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        print(f"DEBUG compras columnas: {list(df.columns)}")

        # Buscar columna nombre producto: "DES ARTICULO", "DESCRIPCION", "ARTICULO", etc.
        prod_col = next((c for c in df.columns if "DES" in c and "ART" in c), None)
        if prod_col is None:
            prod_col = next((c for c in df.columns if "DESCRIPCION" in c or "ARTICULO" in c), None)
        if prod_col is None:
            prod_col = next((c for c in df.columns if "PRODUCTO" in c or "NOMBRE" in c), None)

        # Buscar columna código artículo: "COD ARTICULO"
        cod_col = next((c for c in df.columns if "COD" in c and "ART" in c), None)

        # Buscar columna cantidad: "CANTIDAD", "UNIDADES", "UDS", "BULTOS"
        cant_col = next((c for c in df.columns if "CANTIDAD" in c), None)
        if cant_col is None:
            cant_col = next((c for c in df.columns if "UNIDADES" in c or c == "UDS" or "BULTOS" in c), None)
        if cant_col is None:
            cant_col = next((c for c in df.columns if "UND" in c and "COMPRA" not in c), None)

        print(f"DEBUG compras: prod_col={prod_col}  cod_col={cod_col}  cant_col={cant_col}")

        if prod_col and cant_col:
            _cols = [c for c in [cod_col, prod_col, cant_col] if c]
            df2 = df[_cols].copy()
            df2[prod_col] = df2[prod_col].astype(str).str.strip().str.rstrip(",")
            df2[cant_col] = pd.to_numeric(df2[cant_col], errors="coerce").fillna(0)
            df2 = df2[df2[prod_col].str.len() > 2]
            df2 = df2[~df2[prod_col].str.upper().isin(["DESCRIPCION", "TOTAL", "NAN", ""])]
            resumen = df2.groupby(prod_col)[cant_col].sum()
            datos = {str(k).strip(): float(v) for k, v in resumen.items()}
            # Mapa código → cantidad para matching exacto (código sin prefijo ERE)
            if cod_col:
                cod_res = df2.groupby(cod_col)[cant_col].sum()
                for _c, _v in cod_res.items():
                    _codisys_compras_cod_map[str(_c).strip()] = float(_v)
            print(f"DEBUG compras: {len(datos)} productos, {len(_codisys_compras_cod_map)} códigos")
        else:
            print(f"DEBUG compras: columnas no encontradas. Disponibles: {list(df.columns)}")
    except Exception as e:
        import traceback
        print(f"Error parseando compras: {traceback.format_exc()}")
    return datos


def _parsear_ventas(xlsx_path: Path) -> float:
    """Parsea el Excel de Ventas diarias.
    Devuelve el total de BASE IMPONIBLE (ventas netas sin IVA) sumado para el periodo.
    """
    try:
        df = pd.read_excel(xlsx_path)
        df.columns = [str(c).strip().upper() for c in df.columns]
        print(f"DEBUG ventas columnas: {list(df.columns)}")

        # Buscar columna de importe: preferimos BASE IMPONIBLE (sin IVA), luego TOTAL
        val_col = next((c for c in ["BASE IMPONIBLE", "TOTAL", "IMPORTE"] if c in df.columns), None)
        if val_col is None:
            print(f"DEBUG ventas: no se encontró columna de importe. Disponibles: {list(df.columns)}")
            return 0.0

        # Filtrar solo filas con fecha válida (excluye filas de subtotales/resumen)
        if "FECHA" in df.columns:
            df = df[pd.to_datetime(df["FECHA"], errors="coerce").notna()].copy()

        df[val_col] = pd.to_numeric(df[val_col], errors="coerce").fillna(0)
        total = float(df[val_col].sum())
        print(f"DEBUG ventas: {val_col} total = {total:.2f} ({len(df)} filas)")
        return total
    except Exception:
        import traceback
        print(f"Error parseando ventas: {traceback.format_exc()}")
        return 0.0


async def _fetch_codisys_all(desde: str, hasta: str) -> tuple:
    """
    Entra a Codisys y descarga los tres informes en una sola sesión.
    Devuelve (consumos_dict, compras_dict, ventas_float, error_str).
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return None, None, None, "playwright no instalado"

    # Leer credenciales: 1) st.secrets (cloud) 2) os.environ 3) .env local
    def _leer_cred(clave):
        # Streamlit Secrets
        try:
            if clave in st.secrets:
                return str(st.secrets[clave])
        except Exception:
            pass
        # Variables de entorno (cargadas por _cargar_env)
        v = os.environ.get(clave, "")
        if v:
            return v
        # Fichero .env local
        carpeta = Path(__file__).parent
        for nombre in [".env", ".env.txt", "env.txt"]:
            p = carpeta / nombre
            if p.exists():
                with open(p, encoding="utf-8") as f:
                    for linea in f:
                        linea = linea.strip()
                        if linea.startswith(clave + "="):
                            return linea.split("=", 1)[1].strip()
        return ""

    codisys_user = _leer_cred("CODISYS_USER")
    codisys_pass = _leer_cred("CODISYS_PASS")
    print(f"DEBUG: USER='{codisys_user}' PASS_len={len(codisys_pass)}")

    if not codisys_user or not codisys_pass:
        return None, None, None, f"Credenciales vacías. USER='{codisys_user}'"

    import tempfile
    tmp = Path(tempfile.mkdtemp())
    consumos = {}
    compras  = {}
    ventas   = 0.0

    try:
        async with async_playwright() as p:
            import platform as _platform
            _is_linux = _platform.system() == "Linux"
            _launch_args = ["--disable-dev-shm-usage", "--disable-gpu"]
            if _is_linux:
                _launch_args += ["--no-sandbox", "--disable-setuid-sandbox"]
            browser = await p.chromium.launch(
                headless=_is_linux,
                args=_launch_args,
            )
            context = await browser.new_context(
                accept_downloads=True,
                viewport={"width": 1280, "height": 900}
            )
            page = await context.new_page()

            # Login
            await page.goto(CODISYS_URL)
            await page.wait_for_load_state("networkidle")
            await page.wait_for_timeout(2000)

            # Usuario: probar selector DevExpress primero, luego genérico
            _usr_sel = None
            for _s in ['#UserName_I', 'input[name="UserName"]', 'input[type="text"]']:
                if await page.locator(_s).count() > 0:
                    _usr_sel = _s
                    break
            if _usr_sel:
                await page.fill(_usr_sel, codisys_user)
            await page.wait_for_timeout(400)

            # Contraseña: selector explícito, sin depender del Tab
            _pwd_sel = None
            for _s in ['#Password_I', 'input[name="Password"]', 'input[type="password"]']:
                if await page.locator(_s).count() > 0:
                    _pwd_sel = _s
                    break
            if _pwd_sel:
                # DevExpress oculta el input real — click en contenedor + force=True
                _pwd_container = _pwd_sel.replace('_I', '')  # e.g. #Password_I → #Password
                if await page.locator(_pwd_container).count() > 0:
                    await page.locator(_pwd_container).click()
                    await page.wait_for_timeout(300)
                await page.locator(_pwd_sel).fill(codisys_pass, force=True)
                await page.wait_for_timeout(400)
                await page.keyboard.press('Enter')
            else:
                # Fallback: Tab desde usuario + type
                await page.keyboard.press('Tab')
                await page.wait_for_timeout(300)
                await page.keyboard.type(codisys_pass)
                await page.wait_for_timeout(300)
                await page.keyboard.press('Enter')

            await page.wait_for_load_state("networkidle")
            await page.wait_for_timeout(2000)
            print(f"DEBUG login: URL tras submit = {page.url}")

            # Informe 1: Consumos (buscar en todas las secciones por si ya no está en FAVORITOS)
            path_consumos = await _descargar_excel_reporte(
                page, "Consumo Materias Primas Venta", desde, hasta, tmp, "consumos.xlsx",
                familia="Compras Logistica", tienda_es_tagbox=False, expandir_secciones=True
            )
            if path_consumos:
                import shutil as _shutil
                _shutil.copy(path_consumos, Path(__file__).parent / "debug_consumos.xlsx")
                consumos = _parsear_consumos(path_consumos)

            # Informe 2: Mercancía Recibida (en "Informes de Compra", necesita expandir)
            path_compras = await _descargar_excel_reporte(
                page, "Mercancía Recibida", desde, hasta, tmp, "compras.xlsx",
                familia="Compras Logistica", tienda_es_tagbox=True, expandir_secciones=True
            )
            if path_compras:
                compras = _parsear_compras(path_compras)

            # Informe 3: Ventas diarias (en "Informes de Ventas", necesita expandir)
            path_ventas = await _descargar_excel_reporte(
                page, "Ventas diarias", desde, hasta, tmp, "ventas.xlsx",
                tienda_es_tagbox=True, expandir_secciones=True
            )
            if path_ventas:
                import shutil
                shutil.copy(path_ventas, Path(__file__).parent / "debug_ventas.xlsx")
                ventas = _parsear_ventas(path_ventas)

            await browser.close()

    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Error Codisys: {error_msg}")
        return None, None, None, error_msg

    return (consumos or None), (compras or None), (ventas if ventas else None), None


def buscar_en_codisys(nombre_inv: str, datos_cod: dict, prefijo_debug: str = "") -> float | None:
    """Busca el valor para un producto del inventario usando matching híbrido palabra+secuencia."""
    from difflib import SequenceMatcher
    import re

    _STOP = {"X","DE","Y","A","LA","EL","LOS","LAS","EN","CON","C","UD","UDS","KG","GR","ML","L"}

    def _tokens(s):
        parts = re.findall(r'[A-Z0-9][A-Z0-9,\.]*', s.upper())
        return {t for t in parts if len(t) > 1 and t not in _STOP}

    def _word_score(a, b):
        ta, tb = _tokens(a), _tokens(b)
        if not ta:
            return 0.0
        matches = sum(1 for t in ta if t in tb or any(t in bt or bt in t for bt in tb))
        return matches / len(ta)

    def _seq_score(a, b):
        return SequenceMatcher(None, a.upper(), b.upper()).ratio()

    mejor, score = None, 0.0
    for nombre_cod in datos_cod:
        s = 0.7 * _word_score(nombre_inv, nombre_cod) + 0.3 * _seq_score(nombre_inv, nombre_cod)
        if s > score:
            score = s
            mejor = nombre_cod

    if prefijo_debug:
        umbral_ok = "✅" if score >= 0.50 else "❌"
        print(f"DEBUG {prefijo_debug} '{nombre_inv}' → mejor='{mejor}' score={score:.2f} {umbral_ok}")

    if score >= 0.50 and mejor is not None:
        return float(datos_cod[mejor] or 0)
    return None


def fetch_codisys_sync(desde: str, hasta: str) -> tuple:
    """Devuelve (consumos_dict, compras_dict, ventas_float, error_str)."""
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(_fetch_codisys_all(desde, hasta))
    except Exception as e:
        import traceback
        return None, None, None, traceback.format_exc()


# ═══════════════════════════════════════════════════════════
#  EXPORTAR A EXCEL
# ═══════════════════════════════════════════════════════════

def exportar_excel(resultados: list, fecha: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"INV {fecha.replace('/', '-')}"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font   = Font(bold=True)
    red_fill    = PatternFill("solid", fgColor="FFCCCC")
    green_fill  = PatternFill("solid", fgColor="CCFFCC")
    center      = Alignment(horizontal="center")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:K1")
    ws["A1"] = f"INVENTARIO 100M LAGOH — {fecha}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")

    # Cabeceras
    headers = ["PRODUCTO", "INV. INICIAL", "COMPRAS", "CAJAS", "PAQUETES",
               "SUELTOS", "INV. FINAL", "USADO INV.", "USADO CODISYS",
               "DIFERENCIA", "DIFERENCIA €"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill  = header_fill
        c.font  = header_font
        c.alignment = center
        c.border = border

    # Datos
    total_dif_eur = 0.0
    for i, r in enumerate(resultados, start=4):
        vals = [
            r["nombre"], r["inv_inicial"], r["compras"],
            r["cajas"], r["paquetes"], r["sueltos"],
            r["inv_final"], r["usado_inv"],
            r["usado_cod"] if r["usado_cod"] is not None else "—",
            r["diferencia"] if r["diferencia"] is not None else "—",
            r["dif_eur"] if r["dif_eur"] is not None else "—",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = border
            if col == 1:
                c.font = bold_font
            if isinstance(val, float):
                c.number_format = "#,##0.00"

        # Colorear diferencia
        dif_cell = ws.cell(row=i, column=10)
        if r["diferencia"] is not None:
            if r["diferencia"] < -5:
                dif_cell.fill = red_fill
            elif r["diferencia"] > 5:
                dif_cell.fill = green_fill

        if r["dif_eur"] is not None:
            total_dif_eur += r["dif_eur"]

    # Total
    total_row = len(resultados) + 4
    ws.cell(row=total_row, column=10, value="TOTAL €").font = bold_font
    t = ws.cell(row=total_row, column=11, value=round(total_dif_eur, 2))
    t.font  = bold_font
    t.number_format = "#,##0.00"
    if total_dif_eur < 0:
        t.fill = red_fill
    else:
        t.fill = green_fill

    # Anchos de columna
    col_widths = [44, 12, 10, 8, 10, 8, 10, 11, 13, 12, 13]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════
#  INTERFAZ — Streamlit
# ═══════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Inventario 100M LAGOH",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# CSS visual mejorado
st.markdown("""
<style>
    /* ── Layout base ── */
    .block-container {
        padding-left: 0.8rem !important;
        padding-right: 0.8rem !important;
        padding-top: 0.8rem !important;
        max-width: 1200px;
    }

    /* ── Título principal ── */
    .main h1, div[data-testid="stMarkdownContainer"] h1 {
        font-size: 28px !important;
        font-weight: 800 !important;
        color: #1a1a2e !important;
        letter-spacing: -0.5px;
        margin-bottom: 2px !important;
    }
    h2 { font-size: 16px !important; color: #333 !important; }
    h3 { font-size: 15px !important; color: #444 !important; }

    /* ── Tabla de inventario ── */
    div[data-testid="stDataEditor"] {
        font-size: 13px !important;
        border-radius: 10px !important;
        overflow: hidden !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08) !important;
    }
    div[data-testid="stDataEditor"] td,
    div[data-testid="stDataEditor"] th {
        padding: 3px 5px !important;
    }
    div[data-testid="stDataEditor"] th {
        background: #f0f4ff !important;
        font-weight: 700 !important;
        color: #1a1a2e !important;
    }

    /* ── Inputs numéricos (táctil) ── */
    div[data-testid="stNumberInput"] input {
        font-size: 16px !important;
        height: 42px;
        border-radius: 8px !important;
    }

    /* ── Botones ── */
    .stButton > button {
        min-height: 44px;
        font-size: 15px;
        border-radius: 10px !important;
        font-weight: 600 !important;
        transition: transform 0.1s, box-shadow 0.1s !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #e63946, #c1121f) !important;
        border: none !important;
        color: white !important;
    }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        background: #f8f9fa;
        padding: 4px;
        border-radius: 10px;
    }
    .stTabs [data-baseweb="tab"] {
        font-size: 14px;
        padding: 6px 16px;
        border-radius: 8px !important;
        font-weight: 500;
    }
    .stTabs [aria-selected="true"] {
        background: white !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.1) !important;
        font-weight: 700 !important;
    }

    /* ── Métricas ── */
    div[data-testid="stMetric"] {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 12px 16px !important;
        border: 1px solid #e9ecef;
    }
    div[data-testid="stMetric"] label {
        font-size: 12px !important;
        color: #6c757d !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    div[data-testid="stMetricValue"] {
        font-size: 22px !important;
        font-weight: 800 !important;
        color: #1a1a2e !important;
    }

    /* ── Expanders ── */
    div[data-testid="stExpander"] {
        border-radius: 10px !important;
        border: 1px solid #e9ecef !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    }

    /* ── Alerts / mensajes ── */
    div[data-testid="stAlert"] {
        border-radius: 10px !important;
        font-size: 14px;
    }

    /* ── Caption del perfil ── */
    .profile-badge {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: #f0f4ff;
        border: 1px solid #d0d9ff;
        border-radius: 20px;
        padding: 4px 12px;
        font-size: 13px;
        font-weight: 600;
        color: #1a1a2e;
    }
</style>
""", unsafe_allow_html=True)

# ── CABECERA ─────────────────────────────────────────────
st.markdown(
    "<div style=\'display:flex;align-items:center;justify-content:space-between;padding:24px 0 12px 0\'>"
    "<span style=\'font-size:2rem;font-weight:700\'>📦 Inventario 100M LAGOH</span>"
    "<img src=\'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAT4AAACfCAMAAABX0UX9AAAAkFBMVEX///85OTs1NTc2NjggICMAAAAuLjAyMjQkJCcjIyYbGx4rKy0eHiEoKCv6+voaGh3v7+/MzMzp6enW1tYUFBjd3d3b29v29vaNjY56entycnPk5ORra2y9vb3Q0NCcnJ1LS02CgoNCQkRgYGGurq+Li4xYWFmkpKW6urqxsbKUlJVcXF1JSUs/P0EPDxQAAAo831fzAAARqUlEQVR4nO1d6ZqiOhBlk7AvAioqm6gotvb7v91NwpYg0DNzp2dgvpwfvciiOdSWSlXkOAYGBgYGBgYGBgYGBgYGBgYGBgYGBgYGBgYGBoZ/Eev7JT8B+XkNC3fsHLc88r6m+fyxdLe/+kbbIHajyPXWv3qDGSK62iYQRJ4XBaArq6GhxXtbkQE8BZ4EZNNOy1Gax7A5Z6lhqIoJoUh2eol+w0f/+9heJYHvoBfvp7gfEuApiLJiH88/LEPWbW+oOhCoO5h26P3OgfwVODrNjPJ+xkMS+QEAXUuLH2FwE0LpHryDlMa/f0R/EpYo0CPK+mfc7MGhVyKk24+vdHDz0MZvIGjHX7ajc0DeG5rZZyMzRsdeEWDqhTV+fyu0hcnrgbRgG5iYvdFIB/qEY/8EIOu6LlOk61Iydv/YlKfZh7DL7x7ld+Hgv42FPuGo06KmKPvkfC/Ko6KSDOpg2BEXPm01ZVOBrlfuyaMS/oGhfgdOfcUSn9TxjJY99dGx5IU2QY1o7Aduv5Ooy015F3mbOEpySdVJXs2hi+ePoq+ZPDiSx8/U8MWelQpOpADKT6d/+1Klbm3eukOb4sMgrjYv3zTC74Tzprq046VVWwRvQcqV5E9Qe2bzrlF3vvYcTJD53eXG/RvG983oe10kQ6QZTynVVt6ki+N4UgVFieKXJl9I369e50p73D+8H583ItoyYZBzjjOl2vZQgOtRAiaI5DHariqD4V1ht9eefuvY/gDAwFxCJ5RIIQ/Ib+E0xpESYJlwAbRdNUeU89byZw5MFueMko5J6kF05r2gjkvDoXFMeQfeaD2zRXud5+DVHBl42ouafqzf/QZPTToo9uSxwJiOijuaEupyfdwzXBsdH5HvmeIxOJcyW/GJKN0dFY2Mdj9qQz8t2vb4tG7TTgr9Ad80V9xorWugtPRRVg2MzgtudOgI8uplmnw6muyhE7/lTN4sctSEC1Ea/2pRum2OJkcPPfftV2K6pyfF54mPcm8F9T1XNleExPBO146/lj6Xkipt/E42T6E2nlRAwytTaVHH6F07f7jE8OQkHaCvJH3ClPLxdPgD8PQrpnSX1yadavskwUKmvpZMjNl2CEFr6cupBP6E8uW0CxKu6MWEdsj65KfprKw9ed5sQFomKFmE/rX0UZ5lSvl6rpc30It0NF1ROopLy7Xy0+tPfwM30tz7a9JQNQPYUsZrIu6gtZyvFZV+7Qul7ER1Eb53TVp7syRsd+diKYcqTs1Hi97kBeUNaPIH1k9G7vCFmM4DT8LwgRTSSVDV0LchlXc86uPIuKOm78BxAR1Uguls3rm7wwKMH5mmEk2HHmwTO3ik6xydsSH04mZeDaBjpx2vvJv8QIT89lda5gcy/y4a6OOSVA3SR6aJ3+D26EMMRL2pyLT0EW56PDyfCRJC1ICJHzYpK83cnlK/SYfYC/Gw7esp9Be2b9fRNzk9mQES0k2k1RydlJXm85P2EOvjKLwefcaWsmaYvumVNGIGNHPXW3asAL/5qKT1arLNFumdpalUSM9NYOvfp286cLl2gfcXcvqXsW/yAKLs79uFCXKwrZcgIzdjIuzjDjR9OMj5OfoI8f1CTv8ynJ0iofIw+5QQqzoFQVWrPeRcbCJh8B6kIKr6WaxJ+siwc970QQTRLYppZSTnDa2TJF+cjMY2NH1YfHveeJoU0vfMW3mHcQEDIyX9yaT09TwvDj36EjlJCjltmbnrGERIJxAqrAmVmrR9vRgPp0stOgk4HfeR6YWh2sy5g/z8wqN5lVBeaaoGknYT9az1SSUBp2cd5Px4MQlTAuSikZg2rxKJqMlk8Y7KrtTiQ2exJlWSmuDNf9L2jpQUFdC8SiilPjVpo9Ol/nsk/gV9FNPGbxrSn4RI0td52c5+TZp+aqmoMZ208Zuiz9IGLl8UKAI6L9t5FBGMX0yvtEnx28X8tO2j0oUTq+nzBSUpRhsTul30oY4bP8pzdGVUG0qoJuI+ctVlUQvlDSj1IY13p9QT0wbK9GkdzWRWUchHL6dqOSbOmy8cqnRe3bQHyCzwWM7FIUWXZJmq7jNH35uSfGWBYQudm6Jye532ih8jF1MrkgoZXpeE8x2NR6gCTJH/TSP6o6BnWGTgSph1eUR9yQDFp7OqKZHGG8n2h1TIvUjh6+U7qXwvYdfNQfNPmi6lF55sze5qabDM4EJFh0PluwsAnR6hBIUsw9Kv7xSQS576m3wejM73PN6u5bh9b7Y8ldOeL3ozBGp6/yCzCdLbQgRRvqwPRLwHpT0uX/sxiSvSa+nmApMtCPS6Dh2jOVRQY/L07I3oSxhubHFOLUWCtiOlK7r2ujTBElbIh9BLmdCxV0TxJ5p62TrR4Nmxp43NK8JOgWXJPO6K++1e7B622attFcBUUmzOoKuh+kXcu94yrqylxWZrraN915Ml2OM+M9LJXLas66au93va4Lsqi+3O79X49JMex34BvqCrhiER/c3Km12jsPPfu2/ebrlY9qhcPT+QmX/jryeO7x6lByfzp7tS5dMC57oV4iLt9YwWm94p4XAVeXW2v/sBo+WUynA3PoKozX15bQTW7Wgrer+3SJaE3opDvyu3Hbiu7X5UbqLc1wdbIcZageeOQzayQQN0sAI9SQ1O5huBoiw9zz/jLp177pu9RjDBXFovVg0roy26CL2iqagShAqd46s3yT/LFNWCLoFdX8m/xjbKgGHKaMsYvB2M/5haBpgxDl08IcBxSHYalucoDg7r9WHj3ncnsX9FdPQVXQYAnq1qefHLM6zD7ZLzkqY+j2W0qC42EuerL6kQkialYRG9J5QGtkmy4nOZhVly+3mx+wexDjwvWC812GdgYGBgYPiXwfzz/8DmlD4Zgb+KaLXJXwudXf19bFYlVxyZ9CFsyLS6NTbnPJAHePUbP8/CcFgRxVJJWxVxvtp+WrbbLO1XZt5sjXv3E/hjKifn3JJLdkmigdTfj+Tg157rLSZ34NtdetjUq7rH9dMUeKD4L6XJfmZAkP0qE3dCVWOZf02Ke3TP6tXcxF+9HmUUoGeRrHxJBrJs2te3FaP0Wrhox76SLLDcEGxtQsleQeT1U73do1lnT6+C3NSSuEpVUOeoAi9qx2hzluS6Uwv1tVQVOxsbLcCGQM1l01iZ1UZn0WvvnnVZsXMkz+EZJfxlSQfSAxETZOdzGR7zPcd9AOWlqJrR7Q/p7m1V9cP6EZY+EK4by4men/iNnTgpfn1P7T8AtKpby1UIqkruD4EXnjh/t1F5E8vXHaXwJSxbBjrlaCbcQ79wRxWvC5WfHOc9lSva9sxdcZa1B3KxPZsiAGjs8VWB4oi6vniwd00QXtR6FT34MCRJMoEsYI5uWrs5Vuqjhxqss3sw51U3vDUX/qhbu1oUvylQ0NAYyjR7iLyCilzuJvwLlyRfcb2t7G85GRw5T8Or6NAZZ5odOZINb/jKUmcP0INwgIi609GGr7yMH5HtW5wGD+XVXhq3F3hGzto9qtV58I2bUv6zigW02LhTvdd/H6iSXbQd1FEg4i1CUqGWs9s2Q8VWaKXyrhVyVdRno2rdA9JSG3X7GT4WFifU5R23NdDBdRFzR0wfF0u8igLEWKlKjQLtyq0lcOGiT/QO7qt6TmgzCuVcPUC0ddAHvMvRqFTCnbXpwyUt8tnIuYMvQovlozFUpWHJaYc7ZFCB5P2FWtXsLReoqNQ+/Iw4R0JV96FWDQ+IGiTCbjp9IX1Yq1MB38tTK9EtP13OU6DUbldQtiy1K77MZFDXtcILizVXrMYKMGcGKBmmF/ruSQQ7ABnxlKYo7ZJwB6Oq8Dt/ospmM+IiRUDmLcGUGFvoKLAPLXRcStTSkQsVfYmM91pD9F2gXMKAm4tMJMW71RoVA7ZVuoGqwktxAa9RctE+W8qsJoDuwXVNtPJ406HquCbZDwAFEqlh8WldAOpFK0yosQckY5GJa2yruFuv9jfIlfrKhj54EtpIB9KHekFctChS6LwEf4U59xSJvg0NdStU1aWykiynyuAgIZmxUSEk8q8ofCEqSncy/g/SB2kVAVfqzR4u0GV3+4MpVQND2FSUQvowkfAitM0moi/s7li12jg22V9joO43K60KkGQ7X8p3J0DpUzzUAO87hQzHAOkkKkqhrKDq5uQTt+qpQSYL9RGol+2OGqgsZr0rrKQp63t09CH2oUFoq+xheGTgKGWjkvRplSIXZlWqBuyFVGpA2weVMANQPEoUi8CBEA0FkD4Nxl27TzxukGWgKRnKQNcuyovi9RiFq2MVf6BgvKIP+iUsfUrXX/UQ6r0QYoLT2hEhRFcNl1APF0/PDnBC4aOqNM2BQoTEAXrbru8iBDhcCZ+QA+hGwKndBwKe1jRXoZpTs0y42MdRjoPoq7bGhZqKbZ/SlVmexLqhFUpft5sTdChtp9cheaKA6bUIA7hDGsuVqwhxhdQW8tnGE1YduOVIdqpG1aZTAap7IzwuEi7FTuBDQGSdMH1YsSFXSKigoLV18mbT1oH2o2z3EjsCfLPko/I9rr6U7dM+RMTZEXmLfbWNvNmNNdErdcQ1VgGOKxrOnmK7Q1ehQzHKP9bcQRXhbC+1YLyH7+RJVWk55LetUrXbfWCQmNfPIjawIY2OWi34yLIuoa0DRnZojoF9aC5gX3pXeb2qp19rQMdmfoXrMBLcb2RXc1BUIVRvY1hN0SC2Bg9WSI5TEdN3EkXcWYR6HRqDoLXt1GtfrJ9TAOmV0QcIipozaDS1OTcmWE8Z03ABXe/TVaisUg4fPYocXGCeHM413KC2Sxf0TS9VzfsBlRQZVS2M1LY+GnV8AumLOO8DgIoDRF/tGSyj25nKk2T0nLaJXbcgHds69AzMuysmUnFf1cHm7dbGnMS6kBmGX0DRfB9lRrm9H+2bU6IP30+LLWedVqqkmR+CuUZmv92hGioth2bAtqKA16dmh5jWG1RTwcywBG8UFGHWOaht5ttyasu88MI2z1t9REhaN6EJ7Fn34ccGUrgtL6pd34uvmKcMy18iSaqfVtWOB/5IzKBqW3/wvINV/38S1PrV0DfhrJYrvUMc3W5upX1be+XbZno8XuEbZp8AyLq/aiJL63704Vtd6yKtIF2tVPnzJdnXOaepIB4wMHCB7BPNO7F7vxzr/4PNj2YpQ/lVzz7iyIsGUvSO097qUIbHcHd3qRxe4BHXBOdkVya3WYsegpWvVi//+v8mR7EXpYaxiAjjt+MQuf9bQy58Ws45mc7AwMDAwMDAwMDAwMDAwMDAwMDAwMCwYHxTfeJhpFR5iVsMT+DwTV/1MMLebeZfSvSz6LbPC1wPr+4GwcbD+2ms49jFC3TwFYfjNugHPlivN7rumsPLR+t6LSqI3WbNN1571ZowvZ65XeIWwxPo9rTJXc6KsMwAl1tjmUz23AE1YlgnVNri4LoM3+MOeMF9fXW49R2VqtyfFrfBUrUvuKh6HodVzRtP9RTm/9Zq3aYtPSgxc3jtPI/hAfRHFNYlWS7ipFpPT4P6IEAr3GtkyhJ42hazVhaYX4RVddsgJDdbui+lEPwH0amujTnBqpd7XLUgjuiTcR0A/g47/BqkDx8MXlx7AWIFC1l55qxXVdNR0+c6xMZ2zhJ39p9A1taGWatuJT0/R5WBj47nqvaHO2fctpKq6z3CzjPqdpsMwriolLLclWVzP/zLCbhHV5vw+Leagz2i7mWFmcSiBqWvKj2B0ifUg19Z9R9Q+vCZcVWkhU7UW4LK8/ZVPYaavsjzitZbFEuog/wJkJvWhtj24SLmHNdOcpg+p3YBl6IuD0qRkiMT9omYc5CY2m0/MLR9bt1EXdGHiK0J5ZyFfIX2jyKkyptSONQIi6MAqcuQpBSQ36hqRV3Xtowz4AFcBhicHM7BvsDNLW77iQ7uoSfa4a9wXK+QOhdI4S81bUvdYXgE/RbGuG5Ijj0vjlHn8gH+gYI5LH91hVa8qQ9CRFFjOt0I6yUMCeFpngvZjDcx+hmjpvFqU7bkH6vP+sPfLDf9Xb0MDAwMDAwMDAwMDAwMDAwMfwL/AWSrHIz33eh+AAAAAElFTkSuQmCC\' style=\'height:60px;object-fit:contain\'>"
    "</div>",
    unsafe_allow_html=True
)
st.markdown("<hr style='margin:4px 0 12px 0; border:none; border-top:2px solid #e63946;'>", unsafe_allow_html=True)

# ── Selector de perfil ────────────────────────────────────────────────────
if "perfil" not in st.session_state:
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### 👤 ¿Quién hace el inventario?")
    st.markdown("<br>", unsafe_allow_html=True)
    _pc1, _pc2, _pc3, _pc4 = st.columns(4)
    for _col, _nombre in zip([_pc1, _pc2, _pc3, _pc4], ["Laura", "Eli", "Marta", "David"]):
        with _col:
            _ico = PERFILES[_nombre]["icono"]
            if st.button(f"{_ico}  {_nombre}", width='stretch', type="primary"):
                st.session_state["perfil"] = _nombre
                # Limpiar estado del perfil anterior para re-inicializar tabla
                for _k in ["tabla_base_df", "_last_import_id", "_debug_import",
                           "_editor_version", "ventas_netas_total"]:
                    st.session_state.pop(_k, None)
                st.rerun()
    st.stop()

# Configurar perfil activo (sobreescribe globals para este render)
perfil          = st.session_state["perfil"]
PRODUCTOS       = PERFILES[perfil]["productos"]
TIENE_BIBERONES = PERFILES[perfil]["biberones"]
SIN_CONTEO      = PERFILES[perfil].get("sin_conteo", False)
STATE_FILE      = Path(f"inventario_estado_{perfil.lower()}.json")
HISTORIAL_FILE  = Path(f"inventario_historial_{perfil.lower()}.json")

# Limpiar estado si cambia de perfil
if st.session_state.get("_perfil_activo") != perfil:
    _keep = {"perfil"}
    for _k in list(st.session_state.keys()):
        if _k not in _keep:
            del st.session_state[_k]
    st.session_state["perfil"]         = perfil
    st.session_state["_perfil_activo"] = perfil
    st.rerun()

# Cabecera con perfil activo
_ico_p = PERFILES[perfil]["icono"]
_ch1, _ch2 = st.columns([6, 1])
with _ch1:
    st.markdown(f"<span style='font-size:1.3rem;font-weight:600'>{_ico_p} Perfil: {perfil}</span>", unsafe_allow_html=True)
with _ch2:
    if st.button("Cambiar", key="_btn_cambiar_perfil"):
        for _k in list(st.session_state.keys()):
            del st.session_state[_k]
        st.rerun()

tab_nuevo, tab_historial = st.tabs(["➕ Nuevo inventario", "📋 Historial"])


# ══════════════════════════════════════════════════════════
#  TAB 1 — NUEVO INVENTARIO
# ══════════════════════════════════════════════════════════
with tab_nuevo:
    estado = cargar_estado()

    # Aplicar valores pendientes de Codisys ANTES de renderizar los widgets
    for _prod in PRODUCTOS:
        _np = _prod["nombre"]
        for _prefix in ("consumo_cod", "mercan"):
            _pk = f"_pending_{_prefix}_{_np}"
            if _pk in st.session_state:
                st.session_state[f"{_prefix}_{_np}"] = st.session_state.pop(_pk)
    # Ventas netas: total único (no por producto)
    if "_pending_ventas_netas_total" in st.session_state:
        st.session_state["ventas_netas_total"] = st.session_state.pop("_pending_ventas_netas_total")
    # DEBUG TEMPORAL: ver consumo_cod por producto del perfil activo
    _dbg_lines = []
    for _dbg_p in PRODUCTOS[:10]:
        _dbg_nb = _dbg_p["nombre"]
        _dbg_v = float(st.session_state.get(f"consumo_cod_{_dbg_nb}", 0.0))
        _dbg_lines.append(f"  {_dbg_nb[:35]}={_dbg_v:.1f}")
    st.session_state["_debug_codisys_keys"] = "consumo_cod por prod perfil:\n" + "\n".join(_dbg_lines)

    # ── FECHAS + IMPORTAR en una sola fila compacta ─────────────────────
    _cf1, _cf2 = st.columns(2)
    with _cf1:
        desde_1 = st.date_input("📅 Desde", value=date.today(), key="desde_1", format="DD/MM/YYYY")
    with _cf2:
        hasta_1 = st.date_input("📅 Hasta", value=date.today(), key="hasta_1", format="DD/MM/YYYY")

    desde_str = desde_1.strftime("%d/%m/%Y")
    hasta_str = hasta_1.strftime("%d/%m/%Y")
    fecha_str = hasta_str
    usar_rango2 = False
    desde_2 = hasta_2 = None

    _total_ventas = float(st.session_state.get("ventas_netas_total", 0.0))

    # ── Base df: columnas EDITABLES nunca se sobreescriben aquí (evita reset del editor)
    # Solo se inicializa en arranque; Mercancía/Consumo/Final/Dif se actualizan cada render
    if "tabla_base_df" not in st.session_state:
        _base = []
        for prod in PRODUCTOS:
            nombre  = prod["nombre"]
            ini_def = float(estado.get(nombre, {}).get("final", 0.0) or 0.0)
            _row = {
                "Producto":  nombre[:26], "Ini": ini_def,
                "Cod": COD_CODISYS_MAP.get(nombre, ""),
                "Compras": 0.0, "Cajas": 0.0, "Paquetes": 0.0,
                "Sueltos": 0.0, "Final": 0.0, "Cons Real": 0.0, "Consumo": 0.0,
                "Dif": 0.0, "Dif €": 0.0, "Dif %": 0.0,
            }
            if TIENE_BIBERONES:
                _row["Biberones"] = 0.0
            _base.append(_row)
        st.session_state["tabla_base_df"] = pd.DataFrame(_base)

    # Actualizar Mercancía/Consumo desde Codisys y pre-calcular Final/Dif
    # SIN tocar Ini/Cajas/Paquetes/Sueltos → el delta del editor queda intacto
    _tv_pre = float(st.session_state.get("ventas_netas_total", 0.0))
    _df_b = st.session_state["tabla_base_df"].copy()
    # Asegurar columna Cod (puede faltar en sesiones antiguas)
    if "Cod" not in _df_b.columns:
        _df_b["Cod"] = [COD_CODISYS_MAP.get(p["nombre"], "") for p in PRODUCTOS]
    for _ib, prod in enumerate(PRODUCTOS):
        _nb = prod["nombre"]
        _merc_pre = float(st.session_state.get(f"mercan_{_nb}", 0.0))
        _cons_pre  = float(st.session_state.get(f"consumo_cod_{_nb}", 0.0))
        _df_b.at[_ib, "Compras"] = _merc_pre
        _df_b.at[_ib, "Consumo"]   = _cons_pre
        # Pre-calcular columnas derivadas con los valores que ya están en base_df
        _ini_pre  = float(_df_b.at[_ib, "Ini"])
        _caj_pre  = float(_df_b.at[_ib, "Cajas"])
        _paq_pre  = float(_df_b.at[_ib, "Paquetes"])
        _sue_pre  = float(_df_b.at[_ib, "Sueltos"])
        _bib_pre  = float(_df_b.at[_ib, "Biberones"]) if TIENE_BIBERONES and "Biberones" in _df_b.columns else 0.0
        _coste_pre = prod["precio"] / prod["uds_caja"]
        _fin_pre  = calcular_inv_final(prod, _caj_pre, _paq_pre, _sue_pre, _bib_pre)
        _uso_pre  = _ini_pre + _merc_pre - _fin_pre
        _df_b.at[_ib, "Cons Real"] = round(_uso_pre, 2)
        _dif_pre  = round(_cons_pre - _uso_pre, 2) if _cons_pre else 0.0
        _deur_pre = round(_dif_pre * _coste_pre, 2)
        _dpct_pre = round(_deur_pre / _tv_pre * 100, 2) if _tv_pre else 0.0
        _df_b.at[_ib, "Final"] = round(_fin_pre, 2)
        _df_b.at[_ib, "Dif"]   = _dif_pre
        _df_b.at[_ib, "Dif €"] = _deur_pre
        _df_b.at[_ib, "Dif %"] = _dpct_pre
    # DEBUG TEMPORAL: cuántas claves consumo_cod non-zero hay en session_state
    _dbg_pre_cons = {k: v for k, v in st.session_state.items()
                     if k.startswith("consumo_cod_") and float(v or 0) != 0}
    _dbg_pending  = [k for k in st.session_state if k.startswith("_pending_consumo_cod_")]
    st.session_state["_debug_prerender"] = (
        f"consumo_cod non-zero: {len(_dbg_pre_cons)}\n"
        + "\n".join(f"  {k}={v:.0f}" for k,v in list(_dbg_pre_cons.items())[:5])
        + f"\npending keys: {len(_dbg_pending)}"
        + (f" → {_dbg_pending[:3]}" if _dbg_pending else "")
    )
    st.session_state["tabla_base_df"] = _df_b

    # ── Importar desde Excel / CSV ────────────────────────────────────────
    with st.expander("📂 Importar datos desde Excel / CSV", expanded=False):
        _uploaded = st.file_uploader(
            "Sube tu archivo (.xlsx o .csv)",
            type=["xlsx", "csv"],
            key="upload_inventario",
            label_visibility="collapsed",
        )
        if _uploaded is not None:
            _upload_id = f"{_uploaded.name}_{_uploaded.size}"
            if st.session_state.get("_last_import_id") == _upload_id:
                st.success("✅ Archivo ya importado.")

            else:
                try:
                    # Leer sin cabecera para detectar la fila de títulos
                    if _uploaded.name.endswith(".csv"):
                        _df_raw = pd.read_csv(_uploaded, header=None)
                    else:
                        import openpyxl as _openpyxl
                        _wb = _openpyxl.load_workbook(_uploaded, read_only=True, data_only=True)
                        _active_sheet = _wb.active.title
                        _uploaded.seek(0)
                        _df_raw = pd.read_excel(_uploaded, header=None, sheet_name=_active_sheet)

                    # Buscar fila de cabeceras: debe tener PRODUCTO o ≥2 palabras clave
                    _hrow = 0
                    for _ri in range(min(10, len(_df_raw))):
                        _row_str = " ".join(str(v).upper() for v in _df_raw.iloc[_ri] if pd.notna(v))
                        _kw_n = sum(1 for kw in ["CAJAS", "SUELTOS", "PAQUETES", "INICIAL"] if kw in _row_str)
                        if "PRODUCTO" in _row_str or _kw_n >= 2:
                            _hrow = _ri
                            break

                    _df_imp = _df_raw.iloc[_hrow + 1:].copy()
                    # Renombrar columnas deduplicando las vacías (nan, nan → nan_1, nan_2…)
                    _raw_cols = [str(v).strip().lower() for v in _df_raw.iloc[_hrow]]
                    _seen, _cols_dedup = {}, []
                    for _c in _raw_cols:
                        _seen[_c] = _seen.get(_c, 0) + 1
                        _cols_dedup.append(_c if _seen[_c] == 1 else f"{_c}_{_seen[_c]}")
                    _df_imp.columns = _cols_dedup
                    _df_imp = _df_imp.dropna(how="all").reset_index(drop=True)

                    def _find_col_kw(df, keywords):
                        for col in df.columns:
                            for kw in keywords:
                                if kw in col:
                                    return col
                        return None

                    _col_prod = _find_col_kw(_df_imp, ["producto", "nombre"])
                    # Prioridad: inventario final (cierre periodo anterior = apertura nuevo periodo)
                    # Si no existe, usar inventario inicial
                    _col_ini  = (
                        _find_col_kw(_df_imp, ["inventario inicial", "inicial", "inv ini", "inv_ini"]) or
                        _find_col_kw(_df_imp, ["inventario final", "inv final", "inv_final"])
                    )
                    _col_caj  = _find_col_kw(_df_imp, ["cajas enteras", "cajas"])
                    _col_paq  = _find_col_kw(_df_imp, ["paquetes enteros", "paquetes"])
                    _col_sue  = _find_col_kw(_df_imp, ["kgs/uds sueltos", "sueltos", "suelto"])
                    _col_bib  = _find_col_kw(_df_imp, ["biberones enteros", "biberones"]) if TIENE_BIBERONES else None
                    _col_merc = _find_col_kw(_df_imp, ["compras  codisys", "compras codisys", "compras"])

                    # DEBUG temporal — guardar para mostrar también en "ya importado"
                    _dbg = (f"🔍 col_ini=**{_col_ini}** | col_caj={_col_caj} | "
                            f"col_paq={_col_paq} | col_sue={_col_sue} | "
                            f"num_cols={len(_df_imp.columns)} | "
                            f"cols={list(_df_imp.columns[:12])}")
                    st.session_state["_debug_import"] = _dbg
                    st.info(_dbg)
                    # Mostrar primeras 2 filas raw para ver valores reales
                    st.dataframe(_df_imp.iloc[:3][[c for c in _df_imp.columns if any(x in c for x in ["producto","inventar","compras","codisys","nan"])]])

                    if not any([_col_caj, _col_paq, _col_sue, _col_ini]):
                        st.warning(
                            "No se reconocieron columnas. Encontradas: "
                            + ", ".join(_df_imp.columns.tolist()[:12])
                        )
                    else:
                        from difflib import SequenceMatcher
                        def _sim2(a, b):
                            return SequenceMatcher(None, a.upper(), b.upper()).ratio()

                        _df_b_new = st.session_state["tabla_base_df"].copy()
                        _imported = 0

                        _imp_cols = [
                            ("Ini", _col_ini), ("Cajas", _col_caj),
                            ("Paquetes", _col_paq), ("Sueltos", _col_sue),
                        ]
                        if TIENE_BIBERONES and _col_bib:
                            _imp_cols.append(("Biberones", _col_bib))
                        # Para perfiles sin conteo: importar inventario final → Sueltos
                        # calcular_inv_final(cajas=0,paq=0,suel=inv_final) = inv_final → Final correcto
                        if SIN_CONTEO:
                            _col_fin_xl = _find_col_kw(_df_imp, ["inventario final", "inv final", "inv_final"])
                            if _col_fin_xl:
                                _imp_cols.append(("Sueltos", _col_fin_xl))

                        if _col_prod:
                            # Para cada producto PERFILES, encontrar la fila Excel
                            # con mayor similitud (evita que filas posteriores
                            # sobreescriban el match correcto).
                            _best_rows = {}  # perfiles_idx -> (score, row_xl)
                            for _ir, row_xl in _df_imp.iterrows():
                                nombre_xl = str(row_xl[_col_prod]).strip()
                                if nombre_xl in ("", "nan", "None"):
                                    continue
                                for _pi, _pp in enumerate(PRODUCTOS):
                                    s = _sim2(nombre_xl, _pp["nombre"])
                                    if s >= 0.35:
                                        if _pi not in _best_rows or s > _best_rows[_pi][0]:
                                            _best_rows[_pi] = (s, row_xl)
                            for idx, (_score, row_xl) in _best_rows.items():
                                for _dest_col, _src_col in _imp_cols:
                                    if _src_col:
                                        _v = pd.to_numeric(row_xl[_src_col], errors="coerce")
                                        if not pd.isna(_v):
                                            _df_b_new.at[idx, _dest_col] = float(_v)
                                # Importar compras → session_state (si no hay dato Codisys, prevalece éste)
                                if _col_merc:
                                    _vm = pd.to_numeric(row_xl[_col_merc], errors="coerce")
                                    if not pd.isna(_vm) and float(_vm) != 0:
                                        _nb_idx = PRODUCTOS[idx]["nombre"]
                                        if not st.session_state.get(f"mercan_{_nb_idx}"):
                                            st.session_state[f"mercan_{_nb_idx}"] = float(_vm)
                                _imported += 1
                        else:
                            for _ir, row_xl in _df_imp.iterrows():
                                idx = int(_ir)
                                if idx >= len(PRODUCTOS):
                                    break
                                for _dest_col, _src_col in _imp_cols:
                                    if _src_col:
                                        _v = pd.to_numeric(row_xl[_src_col], errors="coerce")
                                        if not pd.isna(_v):
                                            _df_b_new.at[idx, _dest_col] = float(_v)
                                _imported += 1

                        # DEBUG: mostrar primeros valores importados
                        _debug_vals = []
                        for _di in range(min(4, len(PRODUCTOS))):
                            _debug_vals.append(f"{PRODUCTOS[_di]['nombre'][:20]}={_df_b_new.at[_di,'Ini']:.1f}")
                        _dbg2 = "Ini importados: " + " | ".join(_debug_vals)
                        st.session_state["_debug_import"] += " || " + _dbg2
                        st.info(_dbg2)

                        # Pre-calcular Mercancía, Consumo, Final y columnas Dif
                        # (así el editor ya las muestra correctas sin esperar otro render)
                        _tv = float(st.session_state.get("ventas_netas_total", 0.0))
                        for _ib2, _prod2 in enumerate(PRODUCTOS):
                            _nb2 = _prod2["nombre"]
                            _merc2 = float(st.session_state.get(f"mercan_{_nb2}", 0.0))
                            _cons2 = float(st.session_state.get(f"consumo_cod_{_nb2}", 0.0))
                            _df_b_new.at[_ib2, "Compras"] = _merc2
                            _df_b_new.at[_ib2, "Consumo"]   = _cons2
                            _cajas2 = float(_df_b_new.at[_ib2, "Cajas"])
                            _paq2   = float(_df_b_new.at[_ib2, "Paquetes"])
                            _suel2  = float(_df_b_new.at[_ib2, "Sueltos"])
                            _bib2   = float(_df_b_new.at[_ib2, "Biberones"]) if TIENE_BIBERONES and "Biberones" in _df_b_new.columns else 0.0
                            _ini2   = float(_df_b_new.at[_ib2, "Ini"])
                            _coste2 = _prod2["precio"] / _prod2["uds_caja"]
                            _final2 = calcular_inv_final(_prod2, _cajas2, _paq2, _suel2, _bib2)
                            _usado2 = _ini2 + _merc2 - _final2
                            _dif2   = round(_cons2 - _usado2, 2) if _cons2 else 0.0
                            _deur2  = round(_dif2 * _coste2, 2)
                            _dpct2  = round(_deur2 / _tv * 100, 2) if _tv else 0.0
                            _df_b_new.at[_ib2, "Final"] = round(_final2, 2)
                            _df_b_new.at[_ib2, "Dif"]   = _dif2
                            _df_b_new.at[_ib2, "Dif €"] = _deur2
                            _df_b_new.at[_ib2, "Dif %"] = _dpct2

                        # Guardar valores en base df y forzar nueva versión del editor
                        st.session_state["tabla_base_df"] = _df_b_new
                        st.session_state["_editor_version"] = (
                            st.session_state.get("_editor_version", 0) + 1
                        )
                        st.session_state["_last_import_id"] = _upload_id
                        _cols_ok = [n for n, c in [
                            ("Ini", _col_ini), ("Cajas", _col_caj),
                            ("Paquetes", _col_paq), ("Sueltos", _col_sue),
                        ] if c]
                        st.success(f"✅ {_imported} productos importados · Columnas: {', '.join(_cols_ok)}")
                        st.rerun()

                except Exception as _e:
                    import traceback
                    st.error(f"Error: {_e}")
                    st.code(traceback.format_exc())

    # ── Botones de acción ────────────────────────────────────────────────
    _bc1, _bc2 = st.columns([3, 1])
    with _bc1:
        calcular    = st.button("🔄 Calcular y obtener Codisys", type="primary", width='stretch')
    with _bc2:
        sin_codisys = st.button("💾 Guardar", width='stretch',
                                help="Guarda el inventario sin conectar a Codisys")

    # ── Tabla única: inputs + Codisys + resultados calculados ─────────────
    if TIENE_BIBERONES:
        _COL_ORDER = ["Cod", "Producto", "Ini", "Compras", "Cajas", "Paquetes", "Biberones", "Sueltos",
                      "Final", "Cons Real", "Consumo", "Dif", "Dif €", "Dif %"]
    elif SIN_CONTEO:
        _COL_ORDER = ["Cod", "Producto", "Ini", "Compras",
                      "Final", "Cons Real", "Consumo", "Dif", "Dif €", "Dif %"]
    else:
        _COL_ORDER = ["Cod", "Producto", "Ini", "Compras", "Cajas", "Paquetes", "Sueltos",
                      "Final", "Cons Real", "Consumo", "Dif", "Dif €", "Dif %"]
    _df_ordered = st.session_state["tabla_base_df"].reindex(columns=_COL_ORDER).copy()
    if "Dif €" in _df_ordered.columns:
        _df_ordered["Dif €"] = _df_ordered["Dif €"].apply(
            lambda x: f"{x:.2f} €" if pd.notna(x) else "0.00 €"
        )
    _col_cfg = {
        "Producto":  st.column_config.TextColumn(   "Producto",        width=175,  disabled=True),
        "Cod":       st.column_config.TextColumn(   "Cod Ref",         width=130,  disabled=True),
        "Ini":       st.column_config.NumberColumn( "Ini",       format="%.2f",  min_value=0.0,           width=70),
        "Compras":   st.column_config.NumberColumn( "Compras",   format="%.2f",  disabled=True,           width=80),
        "Cajas":     st.column_config.NumberColumn( "Cajas",     format="%.2f",  min_value=0.0, step=1.0, width=65),
        "Paquetes":  st.column_config.NumberColumn( "Paq",       format="%.2f",  min_value=0.0, step=1.0, width=55),
        "Sueltos":   st.column_config.NumberColumn( "Suel",      format="%.2f",  min_value=0.0,           width=60),
        "Final":     st.column_config.NumberColumn( "Final",     format="%.2f",  disabled=True,           width=70),
        "Cons Real": st.column_config.NumberColumn( "Cons\nReal", format="%.2f",  disabled=True,           width=90),
        "Consumo":   st.column_config.NumberColumn( "Cod\nCons.", format="%.2f",  disabled=True,           width=90),
        "Dif":       st.column_config.NumberColumn( "Dif",       format="%.2f",  disabled=True,           width=65),
        "Dif €":     st.column_config.TextColumn(   "Dif €",                     disabled=True,           width=80),
        "Dif %":     st.column_config.NumberColumn( "Dif %",     format="%.2f",  disabled=True,           width=65),
    }
    if TIENE_BIBERONES:
        _col_cfg["Biberones"] = st.column_config.NumberColumn("Bib", format="%.0f", min_value=0.0, step=1.0, width="small")
    # Sin tabla de conteo — datos vienen del Excel importado
    # Ventas netas se carga desde Codisys (session_state["ventas_netas_total"])
    _editado = _df_ordered

    # ── Calcular resultados y actualizar columnas calculadas en el base df ─
    # (aparecen en el editor en el siguiente render, con un retraso de un ciclo)
    entradas       = {}
    _df_b2         = st.session_state["tabla_base_df"].copy()
    _total_dif_eur = 0.0

    for _i, prod in enumerate(PRODUCTOS):
        nombre   = prod["nombre"]
        row      = _editado.iloc[_i]
        ini      = float(row["Ini"])
        cajas    = float(row["Cajas"])    if "Cajas"    in _editado.columns else 0.0
        paq      = float(row["Paquetes"]) if "Paquetes" in _editado.columns else 0.0
        suel     = (float(row["Final"]) if SIN_CONTEO else
                   (float(row["Sueltos"]) if "Sueltos" in _editado.columns else 0.0))
        bib      = float(row["Biberones"]) if TIENE_BIBERONES and "Biberones" in _editado.columns else 0.0
        merc     = float(row["Compras"])
        cons     = float(row["Consumo"])
        coste_ud = prod["precio"] / prod["uds_caja"]

        inv_final = calcular_inv_final(prod, cajas, paq, suel, bib)
        usado_inv = ini + merc - inv_final

        if cons:
            # Consumo cargado de Codisys: calcular diferencia real
            dif_uds = round(cons - usado_inv, 2)
            dif_eur = round(dif_uds * coste_ud, 2)
        else:
            # Sin consumo Codisys: usar el Dif € pre-calculado del render anterior
            _dif_str = row.get("Dif €", "0.0") if "Dif €" in _editado.columns else "0.0"
            dif_eur  = float(str(_dif_str).replace("€", "").strip() or 0.0)
            dif_uds  = float(row.get("Dif", 0.0)) if "Dif" in _editado.columns else 0.0

        dif_pct = round(dif_eur / _total_ventas * 100, 2) if _total_ventas else 0.0
        _total_dif_eur += dif_eur

        # Solo actualizo columnas calculadas, NUNCA las editables
        _df_b2.at[_i, "Final"]     = round(inv_final, 2)
        _df_b2.at[_i, "Cons Real"] = round(usado_inv, 2)
        _df_b2.at[_i, "Dif"]   = dif_uds
        _df_b2.at[_i, "Dif €"] = dif_eur
        _df_b2.at[_i, "Dif %"] = dif_pct

        entradas[nombre] = {
            "inv_ini": ini, "cajas": cajas, "paquetes": paq, "sueltos": suel,
            "biberones": bib,
            "mercancia": merc, "consumo_cod": cons, "ventas_netas": _total_ventas,
            "inv_final": round(inv_final, 2),
            "cons_real": round(usado_inv, 2),
            "dif_uds": dif_uds, "dif_eur": dif_eur, "dif_pct": dif_pct,
        }

    _total_dif_eur = round(_total_dif_eur, 2)


    # ── Tabla resultado HTML con diseño Excel ──────────────────────────────
    def _build_result_html(df_res, prods, tiene_bib, sin_cont, perfil_nm, ventas, dif_eur_tot, dif_pct_tot, desde="", hasta="", cod_map=None):
        C_GREEN  = "#92D050"
        C_RED    = "#FF0000"
        C_BLUE   = "#00B0F0"
        C_DKBLUE = "#0070C0"
        C_GREY   = "#F2F2F2"
        C_DGREY  = "#D9D9D9"
        TH = (
            "style='background:{bg};color:{fc};font-weight:bold;"
            "text-align:center;padding:5px 8px;font-size:12px;"
            "border:1px solid #bbb;white-space:nowrap'"
        )
        TD = "style='text-align:{al};padding:4px 7px;font-size:12px;border:1px solid #ddd;{extra}'"

        def th(label, bg="#4472C4", fc="white"):
            return f"<th {TH.format(bg=bg, fc=fc)}>{label}</th>"
        def td(val, align="right", extra=""):
            return f"<td {TD.format(al=align, extra=extra)}>{val}</td>"

        # Cabecera con Cod Ref a la izquierda
        cols_html  = th("Cod Ref", "#595959")
        cols_html += th("Producto", "#4472C4")
        cols_html += th("Precio €", "#7030A0")
        cols_html += th("Ini", C_GREEN, "black")
        cols_html += th("Compras", C_RED)
        if not sin_cont:
            cols_html += th("Cajas", C_BLUE, "black")
            cols_html += th("Paq", C_BLUE, "black")
            if tiene_bib:
                cols_html += th("Bib", C_BLUE, "black")
            cols_html += th("Suel", C_BLUE, "black")
        cols_html += th("Final", C_GREEN, "black")
        cols_html += th("Cons Real", "#595959")
        cols_html += th("Cod Cons.", C_RED)
        cols_html += th("Dif", "#595959")
        cols_html += th("Dif €", "#595959")
        cols_html += th("Dif %", "#595959", "white")  # sin borde derecho

        rows_html = ""
        for _ri, _prod in enumerate(prods):
            r = df_res.iloc[_ri]
            def _v(col, fmt="{:.2f}"):
                try: return fmt.format(float(r[col])) if col in df_res.columns else "—"
                except: return "—"
            bg_row = "white" if _ri % 2 == 0 else C_GREY
            red_td = f"color:{C_RED};font-weight:bold;"
            rows_html += f"<tr style='background:{bg_row}'>"
            _cod_ref = (cod_map or {}).get(_prod['nombre'], '')
            rows_html += td(f"<small style='color:#555'>{_cod_ref}</small>", "center")
            rows_html += td(f"<b>{_prod['nombre'][:32]}</b>", "left")
            _precio_color = "#CC0000" if _prod.get('precio', 0) == 0.0 else "#7030A0"
            rows_html += td(f"{_prod.get('precio', 0):.2f} €", extra=f"color:{_precio_color};font-weight:bold;")
            rows_html += td(_v("Ini"))
            rows_html += td(_v("Compras"))
            if not sin_cont:
                rows_html += td(_v("Cajas"),    extra=red_td)
                rows_html += td(_v("Paquetes"), extra=red_td)
                if tiene_bib:
                    rows_html += td(_v("Biberones"), extra=red_td)
                rows_html += td(_v("Sueltos"),  extra=red_td)
            rows_html += td(_v("Final"))
            rows_html += td(_v("Cons Real"))
            rows_html += td(_v("Consumo"))
            rows_html += td(_v("Dif"))
            # Dif € viene como string "X.XX €" → limpiar
            _dif_e_raw = str(r["Dif €"]) if "Dif €" in df_res.columns else "0.00"
            _dif_e_num = float(_dif_e_raw.replace("€","").strip() or 0)
            _dif_color = C_RED if _dif_e_num < 0 else ("#0070C0" if _dif_e_num > 0 else "inherit")
            rows_html += td(f"{_dif_e_num:.2f} €", extra=f"color:{_dif_color};font-weight:bold;")
            rows_html += td(_v("Dif %"))
            rows_html += "</tr>"

        # Fila totales
        _dpc = f"{dif_pct_tot:.2f}%"
        _dc  = C_RED if dif_eur_tot < 0 else ("#0070C0" if dif_eur_tot > 0 else "inherit")
        n_extra = 3 + (1 if tiene_bib else 0) + (0 if sin_cont else 3)
        rows_html += f"<tr style='background:{C_DGREY};font-weight:bold'>"
        rows_html += td("", "center")
        rows_html += td("<b>TOTALES</b>", "left")
        rows_html += "".join(td("—") for _ in range(n_extra))
        rows_html += td("—") * 0  # Final, ConsReal, CodCons, Dif
        rows_html += td("") + td("") + td("") + td("")
        rows_html += td(f"<b>{dif_eur_tot:.2f} €</b>", extra=f"color:{_dc};")
        rows_html += td(f"<b>{_dpc}</b>", extra=f"color:{_dc};")
        rows_html += "</tr>"

        # Footer ventas
        footer = (
            f"<tr style='background:#4472C4;color:white;font-weight:bold'>"
            f"<td colspan='100%' style='text-align:right;padding:5px 10px;font-size:12px;border:1px solid #bbb'>"
            f"Ventas netas: <span style='font-size:14px'>{ventas:,.2f} €</span> &nbsp;│&nbsp; "
            f"Total Dif.: <span style='color:{'#ffcccc' if dif_eur_tot<0 else '#ccffcc'}'>{dif_eur_tot:,.2f} €</span> &nbsp;│&nbsp; "
            f"Dif./Ventas: <span style='color:{'#ffcccc' if dif_pct_tot<0 else '#ccffcc'}'>{dif_pct_tot:.2f}%</span>"
            f"</td></tr>"
        )

        # Header con rango de fechas
        _rango = ""
        if desde or hasta:
            _rango = (
                f"<div style='font-family:Arial,sans-serif;font-size:13px;"
                f"font-weight:bold;color:#333;margin-bottom:6px;padding:6px 10px;"
                f"background:#EEF2FF;border-left:4px solid #4472C4;border-radius:2px'>"
                f"📅 Inventario {perfil_nm} &nbsp;│&nbsp; "
                f"Período: <span style='color:#4472C4'>{desde}</span>"
                f" → <span style='color:#4472C4'>{hasta}</span>"
                f"</div>"
            )
        html = (
            f"<div style='overflow-x:auto;margin-top:12px'>"
            f"{_rango}"
            f"<table style='border-collapse:collapse;width:100%;font-family:Arial,sans-serif'>"
            f"<thead><tr>{cols_html}</tr></thead>"
            f"<tbody>{rows_html}{footer}</tbody>"
            f"</table></div>"
        )
        return html

    st.session_state["tabla_base_df"] = _df_b2
    # Rerun extra para mostrar columnas calculadas (Dif, Final, etc.)
    # Al activarse: guardamos los valores editables en base_df e incrementamos la versión
    # del editor para que el nuevo editor arranque con TODOS los valores correctos visibles.
    if st.session_state.pop("_rerun_for_final", False):
        # Escribir columnas editables actuales en base_df para el nuevo editor
        _df_b3 = st.session_state["tabla_base_df"].copy()
        for _i3, _prod3 in enumerate(PRODUCTOS):
            _row3 = _editado.iloc[_i3]
            _df_b3.at[_i3, "Ini"]      = float(_row3["Ini"])
            # Solo sobreescribir si la columna es visible; si no, preservar valor actual
            if "Cajas"    in _editado.columns: _df_b3.at[_i3, "Cajas"]    = float(_row3["Cajas"])
            if "Paquetes" in _editado.columns: _df_b3.at[_i3, "Paquetes"] = float(_row3["Paquetes"])
            if "Sueltos"  in _editado.columns: _df_b3.at[_i3, "Sueltos"]  = float(_row3["Sueltos"])
            if TIENE_BIBERONES and "Biberones" in _editado.columns:
                _df_b3.at[_i3, "Biberones"] = float(_row3.get("Biberones", 0.0))
        st.session_state["tabla_base_df"] = _df_b3
        # Nuevo key de editor → arranca fresco con todos los datos correctos
        st.session_state["_editor_version"] = st.session_state.get("_editor_version", 0) + 1
        st.rerun()
    _dif_pct_total = round(_total_dif_eur / _total_ventas * 100, 2) if _total_ventas else 0.0
    # Forzar valores en session_state para que los widgets con key muestren el valor actual
    st.session_state["total_dif_eur_disp"] = _total_dif_eur
    st.session_state["total_dif_pct_disp"] = _dif_pct_total
    # Generar tabla HTML estilizada con todos los datos calculados
    _html_res = _build_result_html(
        st.session_state["tabla_base_df"], PRODUCTOS, TIENE_BIBERONES, SIN_CONTEO,
        perfil, _total_ventas, _total_dif_eur, _dif_pct_total,
        desde=desde_str, hasta=hasta_str, cod_map=COD_CODISYS_MAP
    )
    st.session_state["_tabla_resultado_html"] = _html_res

    # ── Tabla resultado estilizada ─────────────────────────
    if "_tabla_resultado_html" in st.session_state:
        st.markdown(st.session_state["_tabla_resultado_html"], unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

    # ── Resultado persistente del último Calcular ───────────
    if "_last_codisys_msg" in st.session_state:
        st.info(st.session_state["_last_codisys_msg"])

    if calcular or sin_codisys:
        # Obtener datos de Codisys
        datos_consumos = None
        datos_compras  = None
        datos_ventas   = None
        if calcular:
            with st.spinner("⏳ Conectando con Codisys (puede tardar 1-2 min)…"):
                c1, p1, v1, err1 = fetch_codisys_sync(desde_str, hasta_str)
                if usar_rango2 and desde_2 and hasta_2:
                    desde_str2 = desde_2.strftime("%d/%m/%Y")
                    hasta_str2 = hasta_2.strftime("%d/%m/%Y")
                    c2, p2, v2, err2 = fetch_codisys_sync(desde_str2, hasta_str2)
                else:
                    c2, p2, v2, err2 = None, None, None, None

            # Combinar ambos rangos sumando valores por producto
            def _merge_dicts(d1, d2):
                if not d1 and not d2:
                    return None
                merged = dict(d1 or {})
                for k, v in (d2 or {}).items():
                    merged[k] = merged.get(k, 0.0) + (v or 0.0)
                return merged or None

            datos_consumos = _merge_dicts(c1, c2)
            datos_compras  = _merge_dicts(p1, p2)
            # ventas es un float (total del periodo), no un dict
            _v1 = v1 or 0.0
            _v2 = v2 or 0.0
            datos_ventas = (_v1 + _v2) if (_v1 or _v2) else None
            error_cod = err1 or err2

            if datos_consumos is None and datos_compras is None and datos_ventas is None:
                _err_short = (error_cod or "sin detalles")[:800]
                st.error(f"❌ Codisys falló: {_err_short}")
            else:
                rangos_txt = f"{desde_str} → {hasta_str}"
                ok_c = f"{len(datos_consumos)} productos" if datos_consumos else "sin datos"
                ok_m = f"{len(datos_compras)} productos" if datos_compras else "sin datos"
                ok_v = f"{datos_ventas:.2f} €" if datos_ventas else "sin datos"
                _cod_msg = f"✅ Codisys ({rangos_txt}) — consumos: {ok_c} | compras: {ok_m} | ventas: {ok_v}"
                st.session_state["_last_codisys_msg"] = _cod_msg
                st.success(_cod_msg)

        # Propagar datos Codisys a session_state para el siguiente render
        if calcular:
            # Limpiar keys viejas con nombres de Codisys (no de PERFILES)
            _perfiles_nombres = {p["nombre"] for p in PRODUCTOS}
            for _k in list(st.session_state.keys()):
                if _k.startswith("consumo_cod_"):
                    if _k[len("consumo_cod_"):] not in _perfiles_nombres:
                        del st.session_state[_k]
                elif _k.startswith("mercan_"):
                    if _k[len("mercan_"):] not in _perfiles_nombres:
                        del st.session_state[_k]
            if datos_consumos:
                # DEBUG: primeros 8 items crudos de Codisys + búsqueda para "MONTADITO"
                _dc_sample = list(datos_consumos.items())[:8]
                _dc_lines = [f"  RAW[{k[:40]}]={v}" for k, v in _dc_sample]
                _dc_lines.append(f"  Total prods Codisys: {len(datos_consumos)}")
                # Buscar específicamente "MONTADITO" en datos_consumos
                _mont = [(k,v) for k,v in datos_consumos.items() if "MONTADITO" in k.upper()]
                _dc_lines.append(f"  MONTADITO en Codisys ({len(_mont)}):")
                for _km, _vm in _mont[:5]:
                    _dc_lines.append(f"    [{_km[:50]}]={_vm}")
                # buscar_en_codisys para PRODUCTOS[0]
                if PRODUCTOS:
                    _nb0 = PRODUCTOS[0]["nombre"]
                    _v0 = buscar_en_codisys(_nb0, datos_consumos)
                    _dc_lines.append(f"  buscar('{_nb0[:40]}')={_v0}")
                # Mapa CODIGO→DESCRIPCION (primeros 5 con MONTADITO)
                _mont_cod = [(cod, des) for cod, des in _codisys_cod_map.items()
                             if "MONTADITO" in des.upper()]
                if _mont_cod:
                    _dc_lines.append(f"  COD→DESC MONTADITO:")
                    for _c, _d in _mont_cod[:5]:
                        _dc_lines.append(f"    {_c} → {_d[:45]}")
                st.session_state["_debug_cod_raw"] = "\n".join(_dc_lines)
                for _prod in PRODUCTOS:
                    _nb = _prod["nombre"]
                    # 1) Match exacto por código (cod_codisys_map.json → Codisys CODIGO)
                    _cod = COD_CODISYS_MAP.get(_nb)
                    if _cod and _cod in _codisys_cod_val_map:
                        _v = _codisys_cod_val_map[_cod]
                        print(f"DEBUG cod-match: {_nb[:30]} ({_cod}) = {_v}")
                    else:
                        # 2) Fallback: fuzzy por nombre
                        _v = buscar_en_codisys(_nb, datos_consumos)
                    if _v is not None:
                        st.session_state[f"_pending_consumo_cod_{_nb}"] = float(_v)
            if datos_compras:
                for _prod in PRODUCTOS:
                    _nb = _prod["nombre"]
                    # 1) Match exacto por código (cod_codisys_map.json → Codisys COD ARTICULO)
                    _ere = COD_CODISYS_MAP.get(_nb, "")
                    _cod_sin_ere = _ere.replace("ERE", "").lstrip("0") if _ere else ""
                    _cod_completo = _ere.replace("ERE", "") if _ere else ""
                    _v_cod = (_codisys_compras_cod_map.get(_cod_completo)
                              or _codisys_compras_cod_map.get(_cod_sin_ere))
                    if _v_cod is not None:
                        _v = _v_cod
                        print(f"DEBUG compras cod-match: {_nb[:30]} ({_ere}) = {_v}")
                    elif _ere:
                        # Tiene código ERE pero no está en compras → no hubo entrega esta semana
                        _v = 0.0
                        print(f"DEBUG compras sin entrega: {_nb[:30]} ({_ere})")
                    else:
                        # Sin código ERE → fallback fuzzy por nombre
                        _v = buscar_en_codisys(_nb, datos_compras)
                    if _v is not None and float(_v) != 0:
                        st.session_state[f"_pending_mercan_{_nb}"] = float(_v)
            if datos_ventas is not None:
                st.session_state["_pending_ventas_netas_total"] = float(datos_ventas)
            st.session_state["_rerun_for_final"] = True
            st.rerun()

        # Guardar estado e historial (sin_codisys o tras calcular sin rerun)
        _res_fin = []
        for _sn2, _sv2 in entradas.items():
            _prod_s2 = next((p for p in PRODUCTOS if p["nombre"] == _sn2), None)
            if _prod_s2:
                _fin2 = calcular_inv_final(
                    _prod_s2, _sv2["cajas"], _sv2.get("paquetes", 0.0),
                    _sv2["sueltos"], _sv2.get("biberones", 0.0)
                )
                _res_fin.append({"nombre": _sn2, "inv_final": _fin2})
        if _res_fin:
            guardar_estado(_res_fin, fecha_str)
        guardar_historial({
            "fecha": fecha_str,
            "desde": desde_str,
            "hasta": hasta_str,
            "perfil": st.session_state.get("perfil_activo", ""),
            "ventas_netas": float(_total_ventas),
            "productos": entradas,
        })
        st.success("✅ Inventario guardado.")


# ══════════════════════════════════════════════════════
#  TAB 2 — HISTORIAL
# ══════════════════════════════════════════════════════
with tab_historial:
    historial = cargar_historial()
    if not historial:
        st.info("Aún no hay inventarios guardados.")
    else:
        st.markdown(f"### 📋 {len(historial)} inventario(s) guardado(s)")
        for _hi, _h in enumerate(reversed(historial)):
            _h_real_idx = len(historial) - 1 - _hi   # índice real en la lista
            _lbl = (f"{_h.get('fecha','?')} — {_h.get('perfil','?')} "
                    f"— ventas: {_h.get('ventas_netas', 0):.2f} €")
            with st.expander(_lbl, expanded=False):
                # Botón borrar
                if st.button(f"🗑️ Borrar este inventario", key=f"del_h_{_h_real_idx}"):
                    borrar_historial_idx(_h_real_idx)
                    st.rerun()

                _prods_h = _h.get("productos", {})
                if _prods_h and isinstance(_prods_h, dict):
                    _ventas_h = _h.get("ventas_netas", 0)
                    _dif_eur_h = sum(v.get("dif_eur", 0) for v in _prods_h.values())
                    _dif_pct_h = round(_dif_eur_h / _ventas_h * 100, 2) if _ventas_h else 0.0
                    _rows_h = []
                    for k, v in _prods_h.items():
                        _rows_h.append({
                            "Cod": "", "Producto": k,
                            "Ini": v.get("inv_ini", 0),
                            "Compras": v.get("mercancia", 0),
                            "Cajas": v.get("cajas", 0),
                            "Paquetes": v.get("paquetes", 0),
                            "Biberones": v.get("biberones", 0),
                            "Sueltos": v.get("sueltos", 0),
                            "Final": v.get("inv_final", 0),
                            "Cons Real": v.get("cons_real", 0),
                            "Consumo": v.get("consumo_cod", 0),
                            "Dif €": v.get("dif_eur", 0),
                            "Dif %": v.get("dif_pct", 0),
                        })
                    _df_h = pd.DataFrame(_rows_h)
                    _perfil_h = _h.get("perfil", "")
                    _tiene_bib_h = any(v.get("biberones", 0) for v in _prods_h.values())
                    _prods_list_h = [{"nombre": k} for k in _prods_h.keys()]
                    _html_h = _build_result_html(
                        _df_h, _prods_list_h, _tiene_bib_h, False,
                        _perfil_h, _ventas_h, _dif_eur_h, _dif_pct_h,
                        _h.get("desde", _h.get("fecha", "")),
                        _h.get("hasta", _h.get("fecha", "")),
                        cod_map=COD_CODISYS_MAP,
                    )
                    st.markdown(_html_h, unsafe_allow_html=True)
